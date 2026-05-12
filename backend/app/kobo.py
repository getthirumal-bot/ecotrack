from __future__ import annotations

import base64
import io
import json
import mimetypes
import os
import time
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import httpx
from openpyxl import Workbook
from urllib.parse import urlencode


class KoboError(RuntimeError):
    pass


@dataclass(frozen=True)
class KoboConfig:
    base_url: str
    token: str

    @staticmethod
    def from_env() -> "KoboConfig":
        base_url = (os.environ.get("KOBO_BASE_URL") or "https://eu.kobotoolbox.org").strip().rstrip("/")
        token = (os.environ.get("KOBO_API_TOKEN") or "").strip()
        if not token:
            raise KoboError("KOBO_API_TOKEN is not set")
        return KoboConfig(base_url=base_url, token=token)


def kobo_prefilled_enketo_link(*, cfg: KoboConfig, asset_uid: str, ecotrack_project_id: str, ecotrack_wbs_id: str) -> str:
    """
    Build a prefilled Enketo link for a new submission.
    Kobo supports prefill via query params matching question names.
    """
    base = f"{cfg.base_url}/#/forms/{asset_uid}/enketo"
    qs = urlencode(
        {
            "ecotrack_project_id": ecotrack_project_id,
            "ecotrack_wbs_id": ecotrack_wbs_id,
        }
    )
    return f"{base}?{qs}"


def _auth_headers(token: str) -> Dict[str, str]:
    return {
        "Authorization": f"Token {token}",
        "Accept": "application/json",
    }


def _odk_single_quoted(s: str) -> str:
    """Safe string literal for XLSForm calculate column (single-quoted XPath string)."""
    return "'" + (s or "").replace("'", "''") + "'"


def _survey_row(
    type_: str,
    name: str = "",
    label: str = "",
    hint: str = "",
    required: str = "",
    relevant: str = "",
    appearance: str = "",
    default: str = "",
    calculation: str = "",
) -> List[str]:
    return [type_, name, label, hint, required, relevant, appearance, default, calculation]


def _xlsx_bytes_for_ecotrack_field_updates_form(
    *,
    form_title: str,
    form_id: str,
    task_choices: List[Tuple[str, str]],
    project_id: str = "",
    project_name: str = "",
    field_supervisor_default: str = "",
    designation_default: str = "",
) -> bytes:
    """
    Build a simple XLSForm (as .xlsx) for Kobo:
    - Ecotrack project id stored via calculate (not shown; Enketo ignores text+hidden reliably)
    - Task picker with minimal appearance (dropdown, not a long radio list)
    - begin_group with relevant: follow-up fields only after a task is chosen
    - field supervisor + designation (defaults from Ecotrack assignee; editable on device)
    - status update, phase, GPS, media, remarks
    """
    wb = Workbook()

    ws_survey = wb.active
    ws_survey.title = "survey"
    ws_choices = wb.create_sheet("choices")
    ws_settings = wb.create_sheet("settings")

    hdr = ["type", "name", "label", "hint", "required", "relevant", "appearance", "default", "calculation"]
    ws_survey.append(hdr)

    if (project_name or "").strip():
        ws_survey.append(
            _survey_row(
                "note",
                "intro_project",
                f"Project: {(project_name or '').strip()}",
                "Activities below are for this project only. Ecotrack records the project automatically.",
                "no",
                "",
                "",
                "",
                "",
            )
        )

    # Stored for sync; never shown in UI (calculate, not text+appearance:hidden)
    ws_survey.append(
        _survey_row(
            "calculate",
            "ecotrack_project_id",
            "",
            "",
            "",
            "",
            "",
            "",
            _odk_single_quoted(project_id or ""),
        )
    )

    ws_survey.append(
        _survey_row(
            "select_one ecotrack_tasks",
            "ecotrack_wbs_id",
            "Select activity",
            "Choose one task, then the rest of the form opens.",
            "yes",
            "",
            "minimal",
            "",
            "",
        )
    )

    rel_after_task = "string-length(${ecotrack_wbs_id}) > 0"
    ws_survey.append(
        _survey_row(
            "begin_group",
            "grp_after_task",
            "Task update",
            "",
            "no",
            rel_after_task,
            "",
            "",
            "",
        )
    )

    ws_survey.append(
        _survey_row(
            "text",
            "field_supervisor",
            "Field supervisor (on site today)",
            "Who is supervising this update? Change if someone else is on site.",
            "no",
            "",
            "",
            (field_supervisor_default or "").strip(),
            "",
        )
    )
    ws_survey.append(
        _survey_row(
            "text",
            "field_designation",
            "Designation / role on site",
            "e.g. Field Manager, Supervisor. Edit if needed.",
            "no",
            "",
            "",
            (designation_default or "").strip(),
            "",
        )
    )

    ws_survey.append(
        _survey_row("text", "submitted_by", "Your name / email (optional)", "If different from your Kobo login.", "no", "", "", "", "")
    )
    ws_survey.append(_survey_row("select_one before_after", "phase", "Phase", "", "yes", "", "", "", ""))
    ws_survey.append(_survey_row("select_one wbs_status", "wbs_status", "Task status", "", "yes", "", "", "", ""))
    ws_survey.append(_survey_row("geopoint", "gps", "GPS location", "", "no", "", "", "", ""))
    ws_survey.append(_survey_row("image", "photo", "Photo (optional)", "", "no", "", "", "", ""))
    ws_survey.append(_survey_row("audio", "audio", "Audio note (optional)", "", "no", "", "", "", ""))
    ws_survey.append(_survey_row("video", "video", "Video (optional)", "", "no", "", "", "", ""))
    ws_survey.append(
        _survey_row(
            "note",
            "instructions",
            "If you are offline, you can save and submit when internet is back.",
            "",
            "no",
            "",
            "",
            "",
            "",
        )
    )
    ws_survey.append(_survey_row("text", "remarks", "Remarks (optional)", "", "no", "", "", "", ""))
    ws_survey.append(_survey_row("end_group", "grp_after_task", "", "", "", "", "", "", ""))

    ws_choices.append(["list_name", "name", "label"])
    for wbs_id, label in task_choices:
        ws_choices.append(["ecotrack_tasks", wbs_id, label])
    ws_choices.append(["before_after", "before", "Before"])
    ws_choices.append(["before_after", "after", "After"])

    ws_choices.append(["wbs_status", "in_progress", "In progress"])
    ws_choices.append(["wbs_status", "pending_approval", "Pending approval"])
    ws_choices.append(["wbs_status", "completed", "Completed"])

    ws_settings.append(["form_title", "form_id"])
    ws_settings.append([form_title, form_id])

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def kobo_import_xlsform(
    *,
    cfg: KoboConfig,
    form_title: str,
    form_id: str,
    task_choices: List[Tuple[str, str]],
    asset_uid: Optional[str] = None,
    project_id: str = "",
    project_name: str = "",
    field_supervisor_default: str = "",
    designation_default: str = "",
) -> str:
    """
    Upload an XLSForm to Kobo and return the import uid.
    Uses the /api/v2/imports/ endpoint (multipart).
    """
    xlsx = _xlsx_bytes_for_ecotrack_field_updates_form(
        form_title=form_title,
        form_id=form_id,
        task_choices=task_choices,
        project_id=project_id,
        project_name=project_name,
        field_supervisor_default=field_supervisor_default,
        designation_default=designation_default,
    )
    files = {
        "file": ("ecotrack_field_updates_v1.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    }
    data = {"library": "false"}
    if asset_uid:
        # Attempt to replace/update existing asset (server supports this in practice even if schema is sparse)
        data["assetUid"] = asset_uid
    url = f"{cfg.base_url}/api/v2/imports/"
    with httpx.Client(timeout=60.0) as client:
        r = client.post(url, headers=_auth_headers(cfg.token), data=data, files=files)
        if r.status_code not in (200, 201):
            raise KoboError(f"Import failed ({r.status_code}): {r.text}")
        payload = r.json()
        uid = payload.get("uid")
        if not uid:
            raise KoboError(f"Import did not return uid: {payload}")
        return uid


def kobo_create_or_update_user_form(
    *,
    cfg: KoboConfig,
    existing_asset_uid: Optional[str],
    user_email: str,
    task_choices: List[Tuple[str, str]],
    project_id: str,
    project_name: str,
    field_supervisor_default: str = "",
    designation_default: str = "",
) -> str:
    """
    Create (or update) a per-user Kobo form that contains only that user's tasks.
    Returns the asset UID (existing or newly created).
    """
    safe = (user_email or "user").lower().replace("@", "_at_").replace(".", "_")
    form_title = f"Ecotrack Field Updates ({user_email})"
    form_id = f"ecotrack_field_updates_{safe}"
    import_uid = kobo_import_xlsform(
        cfg=cfg,
        form_title=form_title,
        form_id=form_id,
        task_choices=task_choices,
        asset_uid=existing_asset_uid,
        project_id=project_id,
        project_name=project_name,
        field_supervisor_default=field_supervisor_default,
        designation_default=designation_default,
    )
    payload = kobo_wait_import(cfg=cfg, import_uid=import_uid, timeout_s=180)
    # If server updated existing asset, it may show under "updated", otherwise "created"
    asset_uid = kobo_extract_created_asset_uid(payload)
    kobo_deploy_form(cfg=cfg, asset_uid=asset_uid)
    return asset_uid


def kobo_wait_import(*, cfg: KoboConfig, import_uid: str, timeout_s: int = 90) -> Dict[str, Any]:
    url = f"{cfg.base_url}/api/v2/imports/{import_uid}/"
    deadline = time.time() + timeout_s
    with httpx.Client(timeout=30.0) as client:
        while True:
            r = client.get(url, headers=_auth_headers(cfg.token))
            if r.status_code != 200:
                raise KoboError(f"Import status failed ({r.status_code}): {r.text}")
            payload = r.json()
            status = (payload.get("status") or "").lower()
            if status in ("complete", "completed", "success"):
                return payload
            if status in ("error", "failed"):
                raise KoboError(f"Import failed: {payload}")
            if time.time() > deadline:
                raise KoboError(f"Import timed out: {payload}")
            time.sleep(2)


def kobo_extract_created_asset_uid(import_payload: Dict[str, Any]) -> str:
    """
    Kobo returns import messages with created/updated assets.
    We attempt to find the first updated/created uid.
    """
    messages = import_payload.get("messages") or {}
    for key in ("created", "updated"):
        arr = messages.get(key) or []
        if isinstance(arr, list) and arr:
            uid = arr[0].get("uid")
            if uid:
                return uid
    # Some servers nest under "detail" in errors; be explicit if not found.
    raise KoboError(f"Could not find created asset uid in import response: {import_payload}")


def kobo_deploy_form(*, cfg: KoboConfig, asset_uid: str) -> Dict[str, Any]:
    url = f"{cfg.base_url}/api/v2/assets/{asset_uid}/deployment/"
    with httpx.Client(timeout=60.0) as client:
        r = client.post(url, headers=_auth_headers(cfg.token), json={"active": True})
        if r.status_code != 200:
            raise KoboError(f"Deploy failed ({r.status_code}): {r.text}")
        return r.json()


def kobo_list_submissions(
    *,
    cfg: KoboConfig,
    asset_uid: str,
    submitted_after: Optional[str] = None,
    limit: int = 300,
) -> List[Dict[str, Any]]:
    """
    List submissions. Uses query on _submission_time if submitted_after provided.
    """
    url = f"{cfg.base_url}/api/v2/assets/{asset_uid}/data/"
    params: Dict[str, Any] = {"start": 0, "limit": min(1000, max(1, limit))}
    if submitted_after:
        params["query"] = json.dumps({"_submission_time": {"$gt": submitted_after}})
    with httpx.Client(timeout=60.0) as client:
        r = client.get(url, headers=_auth_headers(cfg.token), params=params)
        if r.status_code != 200:
            raise KoboError(f"List submissions failed ({r.status_code}): {r.text}")
        payload = r.json()
        results = payload.get("results") or payload  # some servers return raw list
        if isinstance(results, list):
            return results
        if isinstance(results, dict) and "results" in results:
            return results["results"] or []
        if isinstance(results, dict) and "results" not in results and "count" in results:
            return results.get("results") or []
        # PaginatedAssetList style
        if isinstance(payload, dict) and "results" in payload:
            return payload["results"] or []
        return []


def kobo_list_attachments(*, cfg: KoboConfig, asset_uid: str, submission_uid: str) -> List[Dict[str, Any]]:
    url = f"{cfg.base_url}/api/v2/assets/{asset_uid}/data/{submission_uid}/attachments/"
    with httpx.Client(timeout=60.0) as client:
        r = client.get(url, headers=_auth_headers(cfg.token))
        if r.status_code != 200:
            raise KoboError(f"List attachments failed ({r.status_code}): {r.text}")
        payload = r.json()
        results = payload.get("results") if isinstance(payload, dict) else payload
        return results or []


def kobo_download_attachment_base64(
    *,
    cfg: KoboConfig,
    asset_uid: str,
    submission_uid: str,
    attachment_id: int,
) -> Tuple[str, str, str]:
    """
    Returns (filename, content_type, base64).
    """
    # Best-effort endpoint: /attachments/{id}/ (often redirects) OR /attachments/{id}/{suffix}/.
    url = f"{cfg.base_url}/api/v2/assets/{asset_uid}/data/{submission_uid}/attachments/{attachment_id}/"
    with httpx.Client(timeout=120.0, follow_redirects=True) as client:
        r = client.get(url, headers=_auth_headers(cfg.token))
        if r.status_code != 200:
            raise KoboError(f"Download attachment failed ({r.status_code}): {r.text}")
        content = r.content
        content_type = (r.headers.get("content-type") or "").split(";")[0].strip()
        filename = "attachment"
        disp = r.headers.get("content-disposition") or ""
        if "filename=" in disp:
            filename = disp.split("filename=", 1)[1].strip().strip('"')
        if not content_type:
            guessed, _ = mimetypes.guess_type(filename)
            content_type = guessed or "application/octet-stream"
        return filename, content_type, base64.b64encode(content).decode("utf-8")

