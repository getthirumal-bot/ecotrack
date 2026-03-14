from __future__ import annotations

import asyncio
import io
import os
import traceback
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import quote

import base64
from fastapi import Depends, File, FastAPI, Form, HTTPException, Request, Response, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse
from fastapi.responses import Response as RawResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl import load_workbook as openpyxl_load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.sql import func as sqlfunc
from sqlmodel import Session, delete, select

from .auth import create_access_token, get_current_user_optional, hash_password, require_roles, verify_password
from .db import create_db_and_tables, get_session
from .notifications import send_activity_reminders as notify_activity_reminders
from .notifications import send_defect_reminders as notify_defect_reminders
from .seed_data import seed_demo_projects
from .models import (
    BoqItem,
    Defect,
    DefectAttachment,
    DefectAttachmentType,
    DefectSeverity,
    DefectStatus,
    MaterialMaster,
    MaintenanceMonth,
    MaintenanceTask,
    PermissionResource,
    Project,
    ProjectStatus,
    ProjectType,
    Role,
    RolePermission,
    User,
    UserLocation,
    UserProject,
    WbsItem,
    WbsItemType,
    WbsStatus,
)

app = FastAPI(title="Ecotrack")
# Resolve paths from this file so /boq and templates work from any CWD
_BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
templates = Jinja2Templates(directory=os.path.join(_BASE_DIR, "templates"))
app.mount("/static", StaticFiles(directory=os.path.join(_BASE_DIR, "static")), name="static")


@app.exception_handler(HTTPException)
def http_exception_handler(request: Request, exc: HTTPException):
    """When browser gets 401, redirect to login instead of showing JSON."""
    if exc.status_code == 401 and "text/html" in (request.headers.get("accept") or ""):
        return RedirectResponse(url="/login", status_code=302)
    return JSONResponse(status_code=exc.status_code, content={"detail": exc.detail})


def _now() -> datetime:
    return datetime.utcnow()


def _as_money(v: Optional[float]) -> str:
    if v is None:
        return "₹0"
    return f"₹{v:,.0f}"


def _validate_wbs_dates(
    start_date: Optional[str],
    end_date: Optional[str],
    parent: Optional[WbsItem],
) -> Tuple[bool, str]:
    """Validate WBS dates: end >= start; child dates within parent range if parent has dates."""
    start = (start_date or "").strip() or None
    end = (end_date or "").strip() or None
    if not start and not end:
        return True, ""
    if start and end:
        if end < start:
            return False, "End date cannot be earlier than start date."
    if parent:
        p_start = (parent.start_date or "").strip() or None
        p_end = (parent.end_date or "").strip() or None
        if p_start and start and start < p_start:
            return False, "Start date cannot be earlier than parent's start date."
        if p_end and end and end > p_end:
            return False, "End date cannot be later than parent's end date."
    return True, ""


def compute_project_costs(session: Session, project_id: str) -> Dict[str, float]:
    items = session.exec(select(BoqItem).where(BoqItem.project_id == project_id)).all()
    est = sum((i.estimated_quantity or 0.0) * (i.unit_price or 0.0) for i in items)
    act = sum((i.actual_quantity or 0.0) * (i.unit_price or 0.0) for i in items)
    return {"estimated_cost": est, "actual_cost": act, "variance": est - act}


def compute_wbs_progress(session: Session, project_id: str) -> float:
    items = session.exec(select(WbsItem).where(WbsItem.project_id == project_id)).all()
    if not items:
        return 0.0

    by_parent: Dict[Optional[str], List[WbsItem]] = {}
    by_id: Dict[str, WbsItem] = {i.id: i for i in items}
    for i in items:
        by_parent.setdefault(i.parent_id, []).append(i)

    def node_progress(node_id: str) -> float:
        node = by_id[node_id]
        children = by_parent.get(node_id, [])
        if not children:
            return 1.0 if node.status == WbsStatus.completed else 0.0
        weights = [max(0.0, c.weight or 0.0) for c in children]
        denom = sum(weights) or float(len(children))
        total = 0.0
        for c, w in zip(children, weights):
            part = node_progress(c.id)
            total += (w if sum(weights) else 1.0) * part
        return total / denom

    roots = by_parent.get(None, [])
    if not roots:
        return 0.0
    weights = [max(0.0, r.weight or 0.0) for r in roots]
    denom = sum(weights) or float(len(roots))
    total = 0.0
    for r, w in zip(roots, weights):
        total += (w if sum(weights) else 1.0) * node_progress(r.id)
    return round((total / denom) * 100.0, 2)


def build_wbs_tree(
    items: List[WbsItem],
    users_by_id: Dict[str, User],
) -> List[Dict[str, Any]]:
    """Build nested tree from flat WBS items. Each node: item, children, primary_owner_name, secondary_owner_name."""
    by_parent: Dict[Optional[str], List[WbsItem]] = {}
    by_id: Dict[str, WbsItem] = {}
    for i in items:
        by_id[i.id] = i
        by_parent.setdefault(i.parent_id, []).append(i)
    for k in by_parent:
        by_parent[k].sort(key=lambda x: (x.name, x.id))

    def node(item: WbsItem) -> Dict[str, Any]:
        children = by_parent.get(item.id, [])
        u_primary = users_by_id.get(item.primary_owner_id) if item.primary_owner_id and item.primary_owner_id in users_by_id else None
        u_secondary = users_by_id.get(item.secondary_owner_id) if item.secondary_owner_id and item.secondary_owner_id in users_by_id else None
        return {
            "item": item,
            "children": [node(c) for c in children],
            "primary_owner_name": u_primary.name if u_primary else None,
            "secondary_owner_name": u_secondary.name if u_secondary else None,
        }

    roots = by_parent.get(None, [])
    return [node(r) for r in roots]


def build_wbs_parent_options(tree: List[Dict[str, Any]], prefix: str = "") -> List[Dict[str, str]]:
    """Flatten WBS tree into (id, display) for parent dropdown; prefix shows hierarchy (e.g. '  ' or 'Parent > ')."""
    out: List[Dict[str, str]] = []
    for node in tree:
        item = node["item"]
        display = f"{prefix}{item.name} — {item.item_type.value}"
        out.append({"id": item.id, "display": display})
        if node.get("children"):
            out.extend(build_wbs_parent_options(node["children"], prefix=prefix + "  "))
    return out


def build_wbs_dropdown_options(
    wbs_items: List[WbsItem], wbs_by_id: Dict[str, WbsItem]
) -> List[Dict[str, str]]:
    """Build hierarchical labels for WBS dropdown (path root → … → item — type). Use in BOQ, Defects, etc."""
    out: List[Dict[str, str]] = []
    for w in wbs_items:
        path_names = []
        cur: Optional[WbsItem] = w
        while cur is not None:
            path_names.append(cur.name)
            cur = wbs_by_id.get(cur.parent_id) if cur.parent_id else None
        path_names.reverse()
        path = " → ".join(path_names)
        display = f"{path} — {w.item_type.value}"
        out.append({"id": w.id, "display": display})
    out.sort(key=lambda o: o["display"])
    return out


def wbs_path_for_item(w: WbsItem, wbs_by_id: Dict[str, WbsItem]) -> str:
    """Full path for one WBS item using ' -> ' (matches WBS Excel template)."""
    path_names = []
    cur: Optional[WbsItem] = w
    while cur is not None:
        path_names.append(cur.name)
        cur = wbs_by_id.get(cur.parent_id) if cur.parent_id else None
    path_names.reverse()
    return " -> ".join(path_names)


def wbs_path_to_id_map(session: Session, project_id: str) -> Dict[str, str]:
    """Map WBS full path string to wbs_item id for BOQ Excel import."""
    items = session.exec(select(WbsItem).where(WbsItem.project_id == project_id)).all()
    by_id = {i.id: i for i in items}
    return {wbs_path_for_item(i, by_id): i.id for i in items}


def wbs_display_path(wbs_id: Optional[str], wbs_by_id: Dict[str, WbsItem]) -> str:
    """Single WBS item full path for display (e.g. 'Parent → Child → Task')."""
    if not wbs_id:
        return "Unassigned"
    w = wbs_by_id.get(wbs_id)
    if not w:
        return str(wbs_id)
    path_names = []
    cur: Optional[WbsItem] = w
    while cur is not None:
        path_names.append(cur.name)
        cur = wbs_by_id.get(cur.parent_id) if cur.parent_id else None
    path_names.reverse()
    return " → ".join(path_names)


def _defect_list_query(
    project_id: str,
    wbs_filter: str = "",
    severity_filter: str = "",
    status_filter: str = "",
    assigned_filter: str = "",
) -> str:
    """Build query string for /defects list to preserve filters (e.g. after edit/delete)."""
    parts = [f"project_id={quote(project_id, safe='')}"]
    if (wbs_filter or "").strip():
        parts.append(f"wbs_filter={quote((wbs_filter or '').strip(), safe='')}")
    if (severity_filter or "").strip():
        parts.append(f"severity_filter={quote((severity_filter or '').strip(), safe='')}")
    if (status_filter or "").strip():
        parts.append(f"status_filter={quote((status_filter or '').strip(), safe='')}")
    if (assigned_filter or "").strip():
        parts.append(f"assigned_filter={quote((assigned_filter or '').strip(), safe='')}")
    return "?" + "&".join(parts)


def generate_project_exec_summary(
    session: Session,
    project_id: str,
    p: Project,
    costs: Dict[str, float],
    progress: float,
    defects_by_wbs: List[Dict[str, Any]],
    cost_by_wbs: List[Dict[str, Any]],
) -> str:
    """Generate a short executive paragraph: where we stand, pain points, support needs."""
    budget = float(p.budget or 0.0)
    actual = costs.get("actual_cost", 0.0)
    est = costs.get("estimated_cost", 0.0)
    parts = []
    # Status
    if progress >= 90:
        parts.append(f"Project is {progress:.0f}% complete.")
    elif progress >= 50:
        parts.append(f"Project is {progress:.0f}% complete with significant work remaining.")
    else:
        parts.append(f"Project at {progress:.0f}% progress; early stage.")
    # Cost
    if budget > 0:
        over = (actual - budget) / budget
        if over > 0.10:
            parts.append(f"Budget overrun of {over*100:.0f}% (actual {_as_money(actual)} vs budget {_as_money(budget)}) is a critical concern.")
        elif over > 0:
            parts.append(f"Mild budget overrun ({_as_money(actual - budget)} over budget).")
        else:
            parts.append(f"Costs within budget (actual {_as_money(actual)}).")
    # Top cost category
    if cost_by_wbs:
        top = max(cost_by_wbs, key=lambda x: x.get("actual_cost", 0) or 0)
        if top.get("actual_cost"):
            parts.append(f"Highest cost area: {top.get('wbs_name', 'N/A')} ({_as_money(top['actual_cost'])}).")
    # Defects / bottlenecks
    if defects_by_wbs:
        top_defect = max(defects_by_wbs, key=lambda x: x.get("defect_count", 0) or 0)
        cnt = top_defect.get("defect_count", 0)
        if cnt > 0:
            parts.append(f"Most defects linked to: {top_defect.get('wbs_name', 'N/A')} ({cnt} issue(s)); review and support may be needed.")
    # Pain points placeholder
    if budget > 0 and actual > budget:
        parts.append("Recommend cost review and prioritisation.")
    return " ".join(parts) if parts else "No summary data available."


def project_health(*, budget: float, actual_cost: float, progress: float) -> str:
    # Simple heuristic based on NRPT heat-map spec:
    # - Green: on time AND actual <= budget  (we approximate "on time" by progress >= 50% for now)
    # - Yellow: <10% budget overrun OR mild slippage
    # - Red: >10% budget overrun OR critical delay
    if budget <= 0:
        return "green"
    over = (actual_cost - budget) / budget
    if over > 0.10:
        return "red"
    if over > 0.0:
        return "yellow"
    # time/progress proxy
    if progress < 25.0:
        return "yellow"
    return "green"


@app.on_event("startup")
def on_startup() -> None:
    create_db_and_tables()


def clear_all_data(session: Session) -> None:
    """Delete all rows in FK-safe order for a fresh DB."""
    session.exec(delete(DefectAttachment))
    session.exec(delete(Defect))
    session.exec(delete(BoqItem))
    session.exec(delete(WbsItem))
    session.exec(delete(Project))
    session.exec(delete(MaterialMaster))
    session.exec(delete(User))
    session.commit()


def seed_if_empty(session: Session) -> None:
    existing = session.exec(select(User)).first()
    if not existing:
        users = [
            User(email="architect@nrpt.com", name="Architect", role=Role.architect, password_hash=hash_password("password")),
            User(email="owner@nrpt.com", name="Project Owner", role=Role.project_owner, password_hash=hash_password("password")),
            User(email="supervisor@nrpt.com", name="Supervisor", role=Role.supervisor, password_hash=hash_password("password")),
            User(email="field@nrpt.com", name="Field Manager", role=Role.field_manager, password_hash=hash_password("password")),
        ]
        for u in users:
            session.add(u)
        session.commit()
    # Always ensure 10 rich demo projects exist (creates only missing by name)
    seed_demo_projects(session)


@app.get("/seed", response_class=HTMLResponse)
def seed(session: Session = Depends(get_session)) -> str:
    seed_if_empty(session)
    return "Seeded (if empty)."


@app.get("/seed-chukapalli", response_class=HTMLResponse)
def seed_chukapalli(session: Session = Depends(get_session)):
    """One-time: create maintenance project 'Chukapalli' if it does not exist."""
    existing = session.exec(select(Project).where(Project.name == "Chukapalli")).first()
    if existing:
        return HTMLResponse(
            "<!DOCTYPE html><html><head><meta charset='utf-8'><title>Chukapalli</title></head><body>"
            "<h1>Chukapalli</h1><p>Project already exists.</p>"
            "<p><a href='/maintenance'>Open Maintenance</a> · <a href='/login'>Login</a></p></body></html>"
        )
    p = Project(
        name="Chukapalli",
        description="Maintenance project",
        budget=0.0,
        status=ProjectStatus.active,
        project_type="maintenance",
        created_by_user_id=None,
        created_at=_now(),
        updated_at=_now(),
    )
    session.add(p)
    session.commit()
    return HTMLResponse(
        "<!DOCTYPE html><html><head><meta charset='utf-8'><title>Chukapalli</title></head><body>"
        "<h1>Chukapalli created</h1><p>Maintenance project has been created.</p>"
        "<p><a href='/maintenance'>Open Maintenance</a> · <a href='/login'>Login</a></p></body></html>"
    )


def _seed_fresh_impl(session: Session) -> str:
    """Clear all data and repopulate with 10 projects and full supporting data."""
    clear_all_data(session)
    seed_if_empty(session)
    return (
        "<!DOCTYPE html><html><head><meta charset='utf-8'><title>Seed Fresh</title></head><body>"
        "<h1>Fresh DB loaded</h1><p>All data cleared and repopulated with:</p><ul>"
        "<li>4 demo users (architect, owner, supervisor, field)</li>"
        "<li>10 projects with different phases and budgets</li>"
        "<li>Full WBS tree per project (milestones → sub-milestones → tasks)</li>"
        "<li>BOQ/BOM per project (materials at WBS level)</li>"
        "<li>Defects per project</li>"
        "</ul><p><a href='/login'>Go to Login</a></p></body></html>"
    )


@app.get("/seed-fresh", response_class=HTMLResponse)
def seed_fresh(session: Session = Depends(get_session)) -> str:
    """Clear all data and repopulate with 10 projects and full supporting data (users, WBS, BOQ, defects)."""
    return _seed_fresh_impl(session)


@app.get("/seed_fresh", response_class=HTMLResponse)
def seed_fresh_alt(session: Session = Depends(get_session)) -> str:
    """Same as /seed-fresh (alternate URL with underscore)."""
    return _seed_fresh_impl(session)


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        "login.html",
        {"request": request, "demo": True},
    )


@app.post("/login")
def login(
    request: Request,
    email: str = Form(...),
    password: str = Form(...),
    session: Session = Depends(get_session),
):
    seed_if_empty(session)
    user = session.exec(select(User).where(User.email == email.strip().lower())).first()
    if not user or not verify_password(password, user.password_hash):
        return templates.TemplateResponse(
            "login.html",
            {"request": request, "error": "Invalid credentials", "demo": True},
            status_code=400,
        )
    token = create_access_token(user_id=user.id, role=user.role.value)
    r = RedirectResponse(url="/", status_code=303)
    r.set_cookie("nrpt_token", token, httponly=True, samesite="lax")
    return r


@app.get("/logout")
def logout():
    r = RedirectResponse(url="/login", status_code=303)
    r.delete_cookie("nrpt_token")
    return r


def ui_context(session: Session, user: User) -> Dict[str, Any]:
    return {"user": user, "role": user.role.value, "money": _as_money}


@app.get("/", response_class=HTMLResponse)
def home(
    request: Request,
    user: Optional[User] = Depends(get_current_user_optional),
    session: Session = Depends(get_session),
):
    # Not logged in: show login page (so visiting the root URL in a browser works)
    if not user:
        return RedirectResponse("/login", status_code=302)
    # Executives land on dashboard; others land on projects
    if user.role in (Role.architect, Role.project_owner):
        return dashboard(request, user=user, session=session)
    return RedirectResponse("/projects", status_code=303)


def _dashboard_at_a_glance(session: Session) -> Dict[str, Any]:
    """Compute at-a-glance metrics and chart data for the dashboard top section."""
    open_statuses = (DefectStatus.open, DefectStatus.in_progress, DefectStatus.reopened, DefectStatus.pending_approval)
    all_defects = session.exec(select(Defect)).all()
    open_defects = sum(1 for d in all_defects if d.status in open_statuses)
    critical_high = sum(1 for d in all_defects if d.severity in (DefectSeverity.critical, DefectSeverity.high) and d.status in open_statuses)
    pending_wbs = session.exec(select(WbsItem).where(WbsItem.status == WbsStatus.pending_approval)).all()
    pending_defects_approval = session.exec(select(Defect).where(Defect.status == DefectStatus.pending_approval)).all()
    pending_boq = session.exec(select(BoqItem).where(BoqItem.pending_approval == True)).all()
    pending_materials = session.exec(select(MaterialMaster).where(MaterialMaster.pending_approval == True)).all()
    pending_approvals = len(pending_wbs) + len(pending_defects_approval) + len(pending_boq) + len(pending_materials)
    all_wbs = session.exec(select(WbsItem)).all()
    wbs_total = len(all_wbs) or 1
    wbs_completed = sum(1 for w in all_wbs if w.status == WbsStatus.completed)
    wbs_in_progress = sum(1 for w in all_wbs if w.status == WbsStatus.in_progress)
    wbs_pending = sum(1 for w in all_wbs if w.status == WbsStatus.pending)
    wbs_pending_appr = sum(1 for w in all_wbs if w.status == WbsStatus.pending_approval)
    wbs_rejected = sum(1 for w in all_wbs if w.status == WbsStatus.rejected)
    portfolio_progress_pct = round(100.0 * wbs_completed / wbs_total, 1) if wbs_total else 0.0
    total_budget = sum((p.budget or 0.0) for p in session.exec(select(Project)).all())
    total_actual = sum(compute_project_costs(session, p.id)["actual_cost"] for p in session.exec(select(Project)).all())
    budget_used_pct = round(100.0 * total_actual / total_budget, 1) if total_budget and total_budget > 0 else 0.0
    open_defects_list = [d for d in all_defects if d.status in open_statuses]
    defects_critical = sum(1 for d in open_defects_list if getattr(d.severity, "value", str(d.severity)) == "critical")
    defects_high = sum(1 for d in open_defects_list if getattr(d.severity, "value", str(d.severity)) == "high")
    defects_medium = sum(1 for d in open_defects_list if getattr(d.severity, "value", str(d.severity)) == "medium")
    defects_low = sum(1 for d in open_defects_list if getattr(d.severity, "value", str(d.severity)) == "low")
    defects_max = max(defects_critical, defects_high, defects_medium, defects_low, 1)
    defects_open = sum(1 for d in open_defects_list if d.status == DefectStatus.open)
    defects_in_progress = sum(1 for d in open_defects_list if d.status == DefectStatus.in_progress)
    defects_reopened = sum(1 for d in open_defects_list if d.status == DefectStatus.reopened)
    defects_pending_appr = sum(1 for d in open_defects_list if d.status == DefectStatus.pending_approval)
    return {
        "open_defects": open_defects,
        "pending_approvals": pending_approvals,
        "portfolio_progress_pct": portfolio_progress_pct,
        "budget_used_pct": budget_used_pct,
        "critical_high_defects": critical_high,
        "wbs_completed": wbs_completed,
        "wbs_in_progress": wbs_in_progress,
        "wbs_pending": wbs_pending,
        "wbs_pending_appr": wbs_pending_appr,
        "wbs_rejected": wbs_rejected,
        "wbs_total": wbs_total,
        "defects_critical": defects_critical,
        "defects_high": defects_high,
        "defects_medium": defects_medium,
        "defects_low": defects_low,
        "defects_max": defects_max,
        "defects_open": defects_open,
        "defects_in_progress": defects_in_progress,
        "defects_reopened": defects_reopened,
        "defects_pending_appr": defects_pending_appr,
    }


@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(
    request: Request,
    user: User = Depends(require_roles(Role.architect, Role.project_owner)),
    session: Session = Depends(get_session),
):
    projects = session.exec(select(Project).order_by(Project.created_at.desc())).all()
    cards = []
    total_budget = 0.0
    total_actual = 0.0
    for p in projects:
        costs = compute_project_costs(session, p.id)
        progress = compute_wbs_progress(session, p.id)
        health = project_health(budget=p.budget or 0.0, actual_cost=costs["actual_cost"], progress=progress)
        total_budget += p.budget or 0.0
        total_actual += costs["actual_cost"]
        cards.append(
            {
                "id": p.id,
                "name": p.name,
                "budget": p.budget or 0.0,
                "actual": costs["actual_cost"],
                "variance": costs["variance"],
                "progress": progress,
                "health": health,
            }
        )
    at_glance = _dashboard_at_a_glance(session)
    # Budget allocation by project (for pie chart): name, pct, start%, end%, color_index; sort highest first
    budget_allocation = []
    if total_budget and total_budget > 0:
        sorted_by_budget = sorted(cards, key=lambda c: c["budget"] or 0, reverse=True)
        cumulative = 0.0
        for i, c in enumerate(sorted_by_budget):
            pct = 100.0 * (c["budget"] or 0) / total_budget
            end = cumulative + pct
            budget_allocation.append({
                "name": c["name"],
                "pct": round(pct, 1),
                "start": round(cumulative, 1),
                "end": round(end, 1),
                "color_index": i % 10,
            })
            cumulative = end
    ctx = ui_context(session, user)
    ctx.update(
        {
            "request": request,
            "projects": cards,
            "total_projects": len(cards),
            "total_budget": total_budget,
            "total_actual": total_actual,
            "total_variance": total_budget - total_actual,
            "at_glance": at_glance,
            "budget_allocation": budget_allocation,
        }
    )
    return templates.TemplateResponse("dashboard.html", ctx)


@app.get("/maintenance", response_class=HTMLResponse)
def maintenance_dashboard(
    request: Request,
    user: User = Depends(require_roles(Role.architect, Role.project_owner)),
    session: Session = Depends(get_session),
):
    """Separate dashboard for maintenance projects: list projects, monthly view, comparison matrix."""
    maintenance_projects = session.exec(
        select(Project).where(Project.project_type == "maintenance").order_by(Project.name.asc())
    ).all()
    # For each project, get recent months and task counts for summary
    project_summaries = []
    for p in maintenance_projects:
        months = session.exec(
            select(MaintenanceMonth).where(MaintenanceMonth.project_id == p.id).order_by(
                MaintenanceMonth.year.desc(), MaintenanceMonth.month.desc()
            )
        ).limit(12).all()
        month_data = []
        for m in months:
            tasks = session.exec(select(MaintenanceTask).where(MaintenanceTask.maintenance_month_id == m.id)).all()
            done = sum(1 for t in tasks if t.status == "done")
            month_data.append({"month": m, "total": len(tasks), "done": done})
        project_summaries.append({"project": p, "months": month_data})
    ctx = ui_context(session, user)
    ctx.update({"request": request, "project_summaries": project_summaries})
    return templates.TemplateResponse("maintenance.html", ctx)


@app.get("/maintenance/project/{project_id}", response_class=HTMLResponse)
def maintenance_project_detail(
    request: Request,
    project_id: str,
    user: User = Depends(require_roles(Role.architect, Role.project_owner)),
    session: Session = Depends(get_session),
):
    p = session.exec(select(Project).where(Project.id == project_id, Project.project_type == "maintenance")).first()
    if not p:
        raise HTTPException(404, "Maintenance project not found")
    months = session.exec(
        select(MaintenanceMonth).where(MaintenanceMonth.project_id == project_id).order_by(
            MaintenanceMonth.year.desc(), MaintenanceMonth.month.desc()
        )
    ).all()
    # Build comparison matrix: rows = task names (union across months), cols = months
    all_task_names = set()
    month_tasks: Dict[str, List[MaintenanceTask]] = {}
    for m in months:
        tasks = session.exec(select(MaintenanceTask).where(MaintenanceTask.maintenance_month_id == m.id).order_by(MaintenanceTask.sort_order)).all()
        month_tasks[m.id] = list(tasks)
        for t in tasks:
            all_task_names.add(t.name or "(unnamed)")
    task_list = sorted(all_task_names) if all_task_names else []
    now = datetime.utcnow()
    ctx = ui_context(session, user)
    ctx.update({
        "request": request,
        "p": p,
        "months": months,
        "month_tasks": month_tasks,
        "task_list": task_list,
        "current_year": now.year,
        "current_month": now.month,
    })
    return templates.TemplateResponse("maintenance_detail.html", ctx)


@app.post("/maintenance/project/{project_id}/month")
def maintenance_add_month(
    project_id: str,
    year: int = Form(...),
    month: int = Form(...),
    user: User = Depends(require_roles(Role.architect, Role.project_owner)),
    session: Session = Depends(get_session),
):
    p = session.exec(select(Project).where(Project.id == project_id, Project.project_type == "maintenance")).first()
    if not p:
        raise HTTPException(404, "Not found")
    existing = session.exec(
        select(MaintenanceMonth).where(
            MaintenanceMonth.project_id == project_id,
            MaintenanceMonth.year == year,
            MaintenanceMonth.month == month,
        )
    ).first()
    if existing:
        return RedirectResponse(f"/maintenance/project/{project_id}?msg=Month+already+exists", status_code=303)
    m = MaintenanceMonth(project_id=project_id, year=year, month=month)
    session.add(m)
    session.commit()
    session.refresh(m)
    return RedirectResponse(f"/maintenance/project/{project_id}?added=1", status_code=303)


@app.post("/maintenance/project/{project_id}/copy-month")
def maintenance_copy_month(
    project_id: str,
    from_year: int = Form(...),
    from_month: int = Form(...),
    to_year: int = Form(...),
    to_month: int = Form(...),
    user: User = Depends(require_roles(Role.architect, Role.project_owner)),
    session: Session = Depends(get_session),
):
    p = session.exec(select(Project).where(Project.id == project_id, Project.project_type == "maintenance")).first()
    if not p:
        raise HTTPException(404, "Not found")
    from_m = session.exec(
        select(MaintenanceMonth).where(
            MaintenanceMonth.project_id == project_id,
            MaintenanceMonth.year == from_year,
            MaintenanceMonth.month == from_month,
        )
    ).first()
    if not from_m:
        return RedirectResponse(f"/maintenance/project/{project_id}?error=Source+month+not+found", status_code=303)
    to_m = session.exec(
        select(MaintenanceMonth).where(
            MaintenanceMonth.project_id == project_id,
            MaintenanceMonth.year == to_year,
            MaintenanceMonth.month == to_month,
        )
    ).first()
    if not to_m:
        to_m = MaintenanceMonth(project_id=project_id, year=to_year, month=to_month)
        session.add(to_m)
        session.commit()
        session.refresh(to_m)
    source_tasks = session.exec(select(MaintenanceTask).where(MaintenanceTask.maintenance_month_id == from_m.id).order_by(MaintenanceTask.sort_order)).all()
    for i, t in enumerate(source_tasks):
        new_t = MaintenanceTask(maintenance_month_id=to_m.id, name=t.name, status="pending", sort_order=i)
        session.add(new_t)
    session.commit()
    return RedirectResponse(f"/maintenance/project/{project_id}?copied=1", status_code=303)


@app.post("/maintenance/month/{month_id}/task")
def maintenance_add_task(
    month_id: str,
    name: str = Form(...),
    status: str = Form("pending"),
    user: User = Depends(require_roles(Role.architect, Role.project_owner)),
    session: Session = Depends(get_session),
):
    m = session.get(MaintenanceMonth, month_id)
    if not m:
        raise HTTPException(404, "Month not found")
    tasks_in_month = session.exec(select(MaintenanceTask).where(MaintenanceTask.maintenance_month_id == month_id)).all()
    sort_order = max((t.sort_order for t in tasks_in_month), default=-1) + 1
    t = MaintenanceTask(maintenance_month_id=month_id, name=name.strip(), status=status, sort_order=sort_order)
    session.add(t)
    session.commit()
    return RedirectResponse(f"/maintenance/project/{m.project_id}?added_task=1", status_code=303)


@app.post("/maintenance/task/{task_id}/status")
def maintenance_task_set_status(
    task_id: str,
    status: str = Form(...),
    user: User = Depends(require_roles(Role.architect, Role.project_owner)),
    session: Session = Depends(get_session),
):
    t = session.get(MaintenanceTask, task_id)
    if not t:
        raise HTTPException(404, "Task not found")
    t.status = status
    session.add(t)
    session.commit()
    m = session.get(MaintenanceMonth, t.maintenance_month_id)
    return RedirectResponse(f"/maintenance/project/{m.project_id}", status_code=303)


@app.get("/projects", response_class=HTMLResponse)
def projects_page(
    request: Request,
    user: User = Depends(require_roles(Role.architect, Role.project_owner, Role.supervisor, Role.field_manager)),
    session: Session = Depends(get_session),
):
    projects = session.exec(select(Project).order_by(Project.created_at.desc())).all()
    cards = []
    for p in projects:
        costs = compute_project_costs(session, p.id)
        progress = compute_wbs_progress(session, p.id)
        cards.append({"p": p, "progress": progress, "actual": costs["actual_cost"], "variance": costs["variance"]})
    ctx = ui_context(session, user)
    ctx.update({"request": request, "projects": cards})
    return templates.TemplateResponse("projects.html", ctx)


@app.post("/projects/create")
def projects_create(
    name: str = Form(...),
    description: str = Form(""),
    budget: float = Form(0.0),
    status: ProjectStatus = Form(ProjectStatus.active),
    project_type: ProjectType = Form(ProjectType.implementation),
    user: User = Depends(require_roles(Role.architect, Role.project_owner)),
    session: Session = Depends(get_session),
):
    p = Project(
        name=name.strip(),
        description=description.strip(),
        budget=float(budget or 0.0),
        status=status,
        project_type=project_type.value,
        created_by_user_id=user.id,
        created_at=_now(),
        updated_at=_now(),
    )
    session.add(p)
    session.commit()
    return RedirectResponse("/projects", status_code=303)


@app.get("/projects/{project_id}", response_class=HTMLResponse)
def project_detail(
    request: Request,
    project_id: str,
    user: User = Depends(require_roles(Role.architect, Role.project_owner, Role.supervisor, Role.field_manager)),
    session: Session = Depends(get_session),
):
    p = session.exec(select(Project).where(Project.id == project_id)).first()
    if not p:
        raise HTTPException(404, "Project not found")
    costs = compute_project_costs(session, p.id)
    progress = compute_wbs_progress(session, p.id)
    # Cost by category (WBS)
    rollup = compute_boq_rollup_by_wbs(session, p.id)
    wbs_items = session.exec(select(WbsItem).where(WbsItem.project_id == p.id)).all()
    wbs_by_id = {w.id: w for w in wbs_items}
    cost_by_wbs: List[Dict[str, Any]] = []
    for wbs_id, data in rollup.items():
        name = "Unassigned" if wbs_id is None else (wbs_by_id.get(wbs_id).name if wbs_by_id.get(wbs_id) else str(wbs_id))
        cost_by_wbs.append({"wbs_id": wbs_id, "wbs_name": name, "estimated_cost": data["estimated_cost"], "actual_cost": data["actual_cost"], "variance": data["variance"]})
    cost_by_wbs.sort(key=lambda x: (x["actual_cost"] or 0), reverse=True)
    # Defects by WBS / task (bottlenecks) — include actual defects per row so expand shows which to resolve
    defects = session.exec(select(Defect).where(Defect.project_id == p.id)).all()
    defects_per_wbs: Dict[Optional[str], List[Defect]] = {}
    for d in defects:
        wid = d.wbs_item_id
        defects_per_wbs.setdefault(wid, []).append(d)
    defects_by_wbs = []
    for wbs_id, defect_list in defects_per_wbs.items():
        name = "Unassigned" if wbs_id is None else (wbs_by_id.get(wbs_id).name if wbs_by_id.get(wbs_id) else str(wbs_id))
        defects_by_wbs.append({
            "wbs_id": wbs_id,
            "wbs_name": name,
            "defect_count": len(defect_list),
            "defects": defect_list,
        })
    defects_by_wbs.sort(key=lambda x: x["defect_count"], reverse=True)
    # Auto executive summary paragraph
    exec_summary_auto = generate_project_exec_summary(session, p.id, p, costs, progress, defects_by_wbs, cost_by_wbs)
    # Chart data: activities planned vs complete, cost vs actual
    wbs_total = len(wbs_items)
    wbs_completed = sum(1 for w in wbs_items if w.status == WbsStatus.completed)
    chart_activities = {"planned": wbs_total, "complete": wbs_completed}
    chart_cost = {"budget": float(p.budget or 0), "estimated": costs.get("estimated_cost", 0), "actual": costs.get("actual_cost", 0)}
    users = session.exec(select(User).order_by(User.name.asc())).all()
    ctx = ui_context(session, user)
    ctx.update({
        "request": request,
        "p": p,
        "costs": costs,
        "progress": progress,
        "cost_by_wbs": cost_by_wbs,
        "defects_by_wbs": defects_by_wbs,
        "exec_summary_auto": exec_summary_auto,
        "chart_activities": chart_activities,
        "chart_cost": chart_cost,
        "users": users,
    })
    return templates.TemplateResponse("project_detail.html", ctx)


@app.post("/projects/{project_id}/summary")
def project_update_summary(
    project_id: str,
    summary_what_completed: str = Form(""),
    summary_where_we_stand: str = Form(""),
    summary_pain_points: str = Form(""),
    summary_where_heading: str = Form(""),
    user: User = Depends(require_roles(Role.architect, Role.project_owner)),
    session: Session = Depends(get_session),
):
    p = session.exec(select(Project).where(Project.id == project_id)).first()
    if not p:
        raise HTTPException(404, "Project not found")
    p.summary_what_completed = summary_what_completed
    p.summary_where_we_stand = summary_where_we_stand
    p.summary_pain_points = summary_pain_points
    p.summary_where_heading = summary_where_heading
    p.updated_at = _now()
    session.add(p)
    session.commit()
    return RedirectResponse(f"/projects/{project_id}", status_code=303)


@app.get("/projects/{project_id}/activities-filtered", response_class=JSONResponse)
def get_activities_filtered(
    project_id: str,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    assigned_to: Optional[str] = None,
    status_filter: Optional[str] = None,
    show_overdue: Optional[str] = None,
    user: User = Depends(require_roles(Role.architect, Role.project_owner, Role.supervisor, Role.field_manager)),
    session: Session = Depends(get_session),
):
    """Get WBS items filtered by date range, assigned person, status, overdue. Return as tree structure."""
    from datetime import datetime as dt
    today = dt.now().date().isoformat()
    
    query = select(WbsItem).where(WbsItem.project_id == project_id)
    
    # Date range filter
    if from_date:
        from_date = from_date.strip()[:10]
        query = query.where(WbsItem.start_date <= to_date if to_date else WbsItem.end_date >= from_date)
    if to_date:
        to_date = to_date.strip()[:10]
        query = query.where(WbsItem.end_date >= from_date if from_date else WbsItem.end_date <= to_date)
    if from_date and to_date:
        # Activities that overlap with date range
        query = query.where(
            (WbsItem.start_date <= to_date) & (WbsItem.end_date >= from_date)
        )
    
    items = session.exec(query).all()
    
    # Filter by assigned person
    if assigned_to and assigned_to.strip():
        items = [i for i in items if i.primary_owner_id == assigned_to or i.secondary_owner_id == assigned_to]
    
    # Filter by status
    if status_filter and status_filter.strip():
        status_val = status_filter.strip()
        items = [i for i in items if i.status.value == status_val]
    
    # Filter overdue (end_date < today and status not completed) - string comparison for SQLite dates
    if show_overdue == "true":
        items = [i for i in items if i.end_date and str(i.end_date) < today and i.status != WbsStatus.completed]
    
    # Also filter by date range in Python if needed (for better control)
    if from_date and to_date:
        filtered_items = []
        for i in items:
            if i.start_date and i.end_date:
                # Activity overlaps with range if start <= to_date and end >= from_date
                if i.start_date <= to_date and i.end_date >= from_date:
                    filtered_items.append(i)
            elif i.start_date and i.start_date <= to_date:
                filtered_items.append(i)
            elif i.end_date and i.end_date >= from_date:
                filtered_items.append(i)
        items = filtered_items
    
    users_by_id = {u.id: u for u in session.exec(select(User)).all()}
    wbs_tree = build_wbs_tree(items, users_by_id)
    
    # Convert tree to serializable format with overdue info
    def serialize_node(n):
        item = n["item"]
        is_overdue = item.end_date and item.end_date < today and item.status != WbsStatus.completed
        return {
            "id": item.id,
            "name": item.name,
            "status": item.status.value,
            "start_date": item.start_date,
            "end_date": item.end_date,
            "primary_owner": n["primary_owner_name"],
            "secondary_owner": n["secondary_owner_name"],
            "primary_owner_id": item.primary_owner_id,
            "secondary_owner_id": item.secondary_owner_id,
            "is_overdue": is_overdue,
            "children": [serialize_node(c) for c in n["children"]],
        }
    return JSONResponse({
        "items": [serialize_node(n) for n in wbs_tree],
        "from_date": from_date,
        "to_date": to_date,
        "count": len(items),
    })


@app.get("/projects/{project_id}/completion-tracking", response_class=JSONResponse)
def get_completion_tracking(
    project_id: str,
    user: User = Depends(require_roles(Role.architect, Role.project_owner, Role.supervisor, Role.field_manager)),
    session: Session = Depends(get_session),
):
    """Get completion tracking data over time for cricket-style graph (expected vs actual)."""
    from datetime import datetime as dt, timedelta
    items = session.exec(select(WbsItem).where(WbsItem.project_id == project_id)).all()
    today = dt.now().date()
    
    # Calculate expected completion rate over time
    # Assume project started at earliest start_date or today
    start_dates = []
    end_dates = []
    for i in items:
        if i.start_date:
            try:
                start_dates.append(dt.fromisoformat(i.start_date).date())
            except:
                pass
        if i.end_date:
            try:
                end_dates.append(dt.fromisoformat(i.end_date).date())
            except:
                pass
    
    if not start_dates:
        return JSONResponse({"error": "No activities with dates"})
    
    project_start = min(start_dates) if start_dates else today
    project_end = max(end_dates) if end_dates else today + timedelta(days=30)
    total_days = (project_end - project_start).days or 1
    total_tasks = len(items)
    
    # Generate data points for each day/week
    data_points = []
    for day_offset in range(0, total_days + 1, max(1, total_days // 30)):  # ~30 points
        date = project_start + timedelta(days=day_offset)
        date_str = date.isoformat()
        
        # Expected: linear progression (if 30 days, day 15 = 50% complete)
        expected_pct = min(100, (day_offset / total_days) * 100) if total_days > 0 else 0
        expected_count = int((expected_pct / 100) * total_tasks)
        
        # Actual: count completed tasks by this date (compare as strings)
        actual_count = sum(1 for i in items if i.status == WbsStatus.completed and i.end_date and str(i.end_date) <= date_str)
        actual_pct = (actual_count / total_tasks * 100) if total_tasks > 0 else 0
        
        data_points.append({
            "date": date_str,
            "expected_pct": round(expected_pct, 1),
            "actual_pct": round(actual_pct, 1),
            "expected_count": expected_count,
            "actual_count": actual_count,
            "total": total_tasks,
        })
    
    # Calculate run rate
    days_elapsed = (today - project_start).days or 1
    tasks_completed = sum(1 for i in items if i.status == WbsStatus.completed)
    actual_run_rate = (tasks_completed / days_elapsed) if days_elapsed > 0 else 0
    expected_run_rate = (total_tasks / total_days) if total_days > 0 else 0
    required_run_rate = ((total_tasks - tasks_completed) / max(1, (project_end - today).days)) if project_end > today else 0
    
    return JSONResponse({
        "data_points": data_points,
        "run_rate": {
            "actual": round(actual_run_rate, 2),
            "expected": round(expected_run_rate, 2),
            "required": round(required_run_rate, 2),
        },
        "summary": {
            "total_tasks": total_tasks,
            "completed": tasks_completed,
            "remaining": total_tasks - tasks_completed,
            "completion_pct": round((tasks_completed / total_tasks * 100) if total_tasks > 0 else 0, 1),
        },
    })


@app.post("/projects/{project_id}/send-activity-reminders")
def project_send_activity_reminders(
    request: Request,
    project_id: str,
    activity_date: str = Form(...),
    user: User = Depends(require_roles(Role.architect, Role.project_owner)),
    session: Session = Depends(get_session),
):
    """PM: send email and WhatsApp to assignees of selected WBS tasks."""
    p = session.exec(select(Project).where(Project.id == project_id)).first()
    if not p:
        raise HTTPException(404, "Project not found")
    activity_date = (activity_date or "").strip()[:10]
    form_data = request.form()
    selected_wbs_ids = set(form_data.getlist("selected_wbs_ids"))
    if not selected_wbs_ids:
        return RedirectResponse(f"/projects/{project_id}?error=" + quote("Please select at least one activity"), status_code=303)
    items = session.exec(
        select(WbsItem).where(
            WbsItem.project_id == project_id,
            WbsItem.id.in_(selected_wbs_ids),
        )
    ).all()
    owner_ids = set()
    task_lines = []
    for item in items:
        task_lines.append(f"  • {item.name} ({item.status.value})")
        if item.primary_owner_id:
            owner_ids.add(item.primary_owner_id)
        if item.secondary_owner_id:
            owner_ids.add(item.secondary_owner_id)
    if not owner_ids:
        return RedirectResponse(
            f"/projects/{project_id}?msg=" + quote("No assignees found for selected activities"),
            status_code=303,
        )
    users = session.exec(select(User).where(User.id.in_(list(owner_ids)))).all()
    task_summary = "\n".join(task_lines) if task_lines else "No task names"
    notify_activity_reminders(list(users), p.name, activity_date, task_summary)
    return RedirectResponse(
        f"/projects/{project_id}?sent_activity=1&count={len(users)}&date={activity_date}",
        status_code=303,
    )


@app.post("/defects/send-reminders")
def defects_send_reminders(
    request: Request,
    project_id: str = Form(...),
    user: User = Depends(require_roles(Role.architect, Role.project_owner)),
    session: Session = Depends(get_session),
):
    """PM: send email and WhatsApp to assignees of selected defects."""
    p = session.exec(select(Project).where(Project.id == project_id)).first()
    if not p:
        raise HTTPException(404, "Project not found")
    form_data = request.form()
    selected_defect_ids = set(form_data.getlist("selected_defect_ids"))
    if not selected_defect_ids:
        return RedirectResponse(f"/defects?project_id={project_id}&error=" + quote("Please select at least one defect"), status_code=303)
    defects = session.exec(
        select(Defect).where(Defect.id.in_(selected_defect_ids), Defect.project_id == project_id)
    ).all()
    assignee_ids = {d.assigned_to_user_id for d in defects if d.assigned_to_user_id}
    if not assignee_ids:
        return RedirectResponse(
            f"/defects?project_id={project_id}&msg=" + quote("No assignees found for selected defects"),
            status_code=303,
        )
    users = session.exec(select(User).where(User.id.in_(list(assignee_ids)))).all()
    defect_lines = [f"  • {d.location}: {d.description[:60]}… ({d.status.value})" for d in defects if d.assigned_to_user_id]
    defect_summary = "\n".join(defect_lines) if defect_lines else "Selected defects."
    notify_defect_reminders(list(users), p.name, defect_summary)
    return RedirectResponse(
        f"/defects?project_id={project_id}&sent_defect=1&count={len(users)}",
        status_code=303,
    )


@app.get("/wbs", response_class=HTMLResponse)
def wbs_page(
    request: Request,
    project_id: Optional[str] = None,
    user: User = Depends(require_roles(Role.architect, Role.project_owner, Role.supervisor, Role.field_manager)),
    session: Session = Depends(get_session),
):
    projects = session.exec(select(Project).order_by(Project.name.asc())).all()
    if not project_id and projects:
        project_id = projects[0].id
    items = session.exec(select(WbsItem).where(WbsItem.project_id == (project_id or ""))).all() if project_id else []
    users = session.exec(select(User).order_by(User.name.asc())).all()
    users_by_id = {u.id: u for u in users}
    wbs_tree = build_wbs_tree(items, users_by_id) if items else []
    wbs_parent_options = build_wbs_parent_options(wbs_tree) if wbs_tree else []
    ctx = ui_context(session, user)
    ctx.update({
        "request": request,
        "projects": projects,
        "project_id": project_id,
        "items": items,
        "wbs_tree": wbs_tree,
        "wbs_parent_options": wbs_parent_options,
        "users": users,
    })
    return templates.TemplateResponse("wbs.html", ctx)


@app.post("/wbs/create")
def wbs_create(
    project_id: str = Form(...),
    parent_id: str = Form(""),
    name: str = Form(...),
    item_type: WbsItemType = Form(WbsItemType.task),
    weight: float = Form(0.0),
    status: WbsStatus = Form(WbsStatus.pending),
    start_date: str = Form(""),
    end_date: str = Form(""),
    primary_owner_id: str = Form(""),
    secondary_owner_id: str = Form(""),
    user: User = Depends(require_roles(Role.architect, Role.project_owner, Role.supervisor, Role.field_manager)),
    session: Session = Depends(get_session),
):
    parent = session.exec(select(WbsItem).where(WbsItem.id == parent_id)).first() if parent_id else None
    ok, err = _validate_wbs_dates(start_date, end_date, parent)
    if not ok:
        return RedirectResponse(f"/wbs?project_id={project_id}&error={quote(err)}", status_code=303)
    # Supervisor/Field create goes to approval queue
    effective_status = status
    if user.role in (Role.supervisor, Role.field_manager):
        effective_status = WbsStatus.pending_approval
    item = WbsItem(
        project_id=project_id,
        parent_id=parent_id or None,
        name=name.strip(),
        item_type=item_type,
        weight=float(weight or 0.0),
        status=effective_status,
        start_date=start_date.strip() or None,
        end_date=end_date.strip() or None,
        primary_owner_id=primary_owner_id or None,
        secondary_owner_id=secondary_owner_id or None,
        created_at=_now(),
        updated_at=_now(),
    )
    session.add(item)
    session.commit()
    return RedirectResponse(f"/wbs?project_id={project_id}", status_code=303)


@app.post("/wbs/{item_id}/status")
def wbs_update_status(
    item_id: str,
    project_id: str = Form(...),
    status: WbsStatus = Form(...),
    open_ids: str = Form(""),
    user: User = Depends(require_roles(Role.architect, Role.project_owner, Role.supervisor, Role.field_manager)),
    session: Session = Depends(get_session),
):
    item = session.exec(select(WbsItem).where(WbsItem.id == item_id)).first()
    if not item:
        raise HTTPException(404, "WBS item not found")
    if user.role == Role.field_manager and status == WbsStatus.rejected:
        raise HTTPException(403, "Field manager cannot reject")
    # Architect/PO can set status directly; Supervisor/Field completion goes to approval queue
    if status == WbsStatus.completed and user.role in (Role.field_manager, Role.supervisor):
        item.status = WbsStatus.pending_approval
    else:
        item.status = status
    item.updated_at = _now()
    session.add(item)
    session.commit()
    url = f"/wbs?project_id={project_id}"
    if (open_ids or "").strip():
        url += f"&open={quote((open_ids or '').strip())}"
    return RedirectResponse(url, status_code=303)


@app.post("/wbs/{item_id}/update")
def wbs_update_full(
    item_id: str,
    project_id: str = Form(...),
    name: str = Form(...),
    item_type: WbsItemType = Form(...),
    weight: float = Form(0.0),
    status: WbsStatus = Form(...),
    start_date: str = Form(""),
    end_date: str = Form(""),
    primary_owner_id: str = Form(""),
    secondary_owner_id: str = Form(""),
    open_ids: str = Form(""),
    user: User = Depends(require_roles(Role.architect, Role.project_owner)),
    session: Session = Depends(get_session),
):
    """Architect and Project Owner only: full inline edit (dates, resources, weight, name, type, status)."""
    item = session.exec(select(WbsItem).where(WbsItem.id == item_id)).first()
    if not item:
        raise HTTPException(404, "WBS item not found")
    parent = session.exec(select(WbsItem).where(WbsItem.id == item.parent_id)).first() if item.parent_id else None
    ok, err = _validate_wbs_dates(start_date, end_date, parent)
    if not ok:
        url = f"/wbs?project_id={project_id}&error={quote(err)}"
        if (open_ids or "").strip():
            url += f"&open={quote((open_ids or '').strip())}"
        return RedirectResponse(url, status_code=303)
    item.name = (name or item.name).strip() or item.name
    item.item_type = item_type
    item.weight = float(weight or 0.0)
    item.status = status
    item.start_date = (start_date or "").strip() or None
    item.end_date = (end_date or "").strip() or None
    item.primary_owner_id = (primary_owner_id or "").strip() or None
    item.secondary_owner_id = (secondary_owner_id or "").strip() or None
    item.updated_at = _now()
    session.add(item)
    session.commit()
    url = f"/wbs?project_id={project_id}"
    if (open_ids or "").strip():
        url += f"&open={quote((open_ids or '').strip())}"
    return RedirectResponse(url, status_code=303)


def _wbs_excel_instructions() -> List[str]:
    return [
        "WBS UPLOAD TEMPLATE - INSTRUCTIONS",
        "",
        "COLOR LEGEND (WBS_Data rows):",
        "- milestone: green",
        "- sub_milestone: yellow",
        "- task: blue",
        "",
        "BRANCHING STRUCTURE (like MS Project):",
        "WBS is a tree: each item can have a parent and children. Use Parent_Path to define branches.",
        "  - Empty Parent_Path = top-level (root) branch.",
        "  - Parent_Path = 'Milestone A' = item is a child of 'Milestone A' (one level down).",
        "  - Parent_Path = 'Milestone A -> Sub B' = item is under Sub B, which is under Milestone A (two levels down).",
        "List parent rows BEFORE their children so the branch is built correctly.",
        "",
        "HOW TO USE THIS FILE:",
        "1. Fill in the 'WBS_Data' sheet with your WBS items (see column headers and sample rows).",
        "2. Save the file as Excel (.xlsx). Do not change sheet names or column headers.",
        "3. On the WBS screen, select your project and click 'Upload' to load this file.",
        "4. Rows with an empty 'Name' are skipped. Parent items must appear before their children.",
        "",
        "COLUMN DESCRIPTIONS:",
        "- Parent_Path: Branch hierarchy using ' -> ' (space, arrow, space). Leave EMPTY for root/top-level items.",
        "  Examples: '' = root | 'Earthwork & Drainage' = child of root | 'Earthwork & Drainage -> Grading and levelling' = level 2.",
        "- Name: Full name of the WBS item (required).",
        "- Type: Exactly one of: milestone | sub_milestone | task",
        "- Weight: Number (e.g. 25 for 25% weight).",
        "- Status: Exactly one of: pending | in_progress | pending_approval | completed | rejected",
        "- Start_Date, End_Date: Format YYYY-MM-DD (e.g. 2025-01-15). Leave empty if not set.",
        "- Primary_Owner, Secondary_Owner: User email or full name as in Ecotrack. Leave empty if unassigned.",
        "",
        "SAMPLE DATA: The first few rows in WBS_Data are examples. Replace or add rows as needed.",
    ]


def build_wbs_excel_template(session: Session, project_id: str) -> bytes:
    """Build Excel with Instructions sheet and WBS_Data sheet (headers + sample rows + colors)."""
    wb = Workbook()
    # Instructions sheet
    ws_inst = wb.active
    ws_inst.title = "Instructions"
    for row_idx, line in enumerate(_wbs_excel_instructions(), start=1):
        ws_inst.cell(row=row_idx, column=1, value=line)
    ws_inst.column_dimensions["A"].width = 100
    # WBS_Data sheet
    ws_data = wb.create_sheet("WBS_Data", 1)
    headers = ["Parent_Path", "Name", "Type", "Weight", "Status", "Start_Date", "End_Date", "Primary_Owner", "Secondary_Owner"]
    for col, h in enumerate(headers, start=1):
        c = ws_data.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F2A44")
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Make it readable (non-MS-Project users): more sample branches and varied statuses/owners
    sample = [
        # Branch 1: Earthwork & Drainage
        ("", "Earthwork & Drainage", "milestone", 25, "in_progress", "2025-01-01", "2025-03-31", "field@nrpt.com", "supervisor@nrpt.com"),
        ("Earthwork & Drainage", "Grading and levelling", "sub_milestone", 60, "in_progress", "2025-01-10", "2025-02-28", "field@nrpt.com", "supervisor@nrpt.com"),
        ("Earthwork & Drainage -> Grading and levelling", "Fine grading and compaction", "task", 50, "completed", "2025-01-15", "2025-01-25", "field@nrpt.com", ""),
        ("Earthwork & Drainage -> Grading and levelling", "Rough grading", "task", 50, "in_progress", "2025-01-26", "2025-02-10", "field@nrpt.com", ""),
        ("Earthwork & Drainage", "Site clearance and stripping", "sub_milestone", 40, "completed", "2025-01-02", "2025-01-09", "field@nrpt.com", ""),
        ("Earthwork & Drainage -> Site clearance and stripping", "Remove debris & dispose", "task", 50, "completed", "2025-01-02", "2025-01-05", "field@nrpt.com", ""),
        ("Earthwork & Drainage -> Site clearance and stripping", "Marking & survey", "task", 50, "completed", "2025-01-06", "2025-01-09", "field@nrpt.com", ""),

        # Branch 2: Hardscape & Finishing
        ("", "Hardscape & Finishing", "milestone", 25, "pending", "2025-03-01", "2025-05-15", "field@nrpt.com", "supervisor@nrpt.com"),
        ("Hardscape & Finishing", "Boundary wall (10ft)", "sub_milestone", 50, "pending", "2025-03-05", "2025-04-10", "field@nrpt.com", ""),
        ("Hardscape & Finishing -> Boundary wall (10ft)", "Brickwork", "task", 40, "pending", "2025-03-05", "2025-03-25", "field@nrpt.com", ""),
        ("Hardscape & Finishing -> Boundary wall (10ft)", "Plaster & finish", "task", 30, "pending", "2025-03-26", "2025-04-05", "field@nrpt.com", ""),
        ("Hardscape & Finishing -> Boundary wall (10ft)", "Paint / coating", "task", 30, "pending", "2025-04-06", "2025-04-10", "field@nrpt.com", ""),
        ("Hardscape & Finishing", "Paving & kerbs", "sub_milestone", 50, "pending", "2025-04-01", "2025-05-10", "field@nrpt.com", ""),
        ("Hardscape & Finishing -> Paving & kerbs", "Sub-base preparation", "task", 35, "pending", "2025-04-01", "2025-04-12", "field@nrpt.com", ""),
        ("Hardscape & Finishing -> Paving & kerbs", "Paver laying", "task", 45, "pending", "2025-04-13", "2025-05-02", "field@nrpt.com", ""),
        ("Hardscape & Finishing -> Paving & kerbs", "Joint sand & compaction", "task", 20, "pending", "2025-05-03", "2025-05-10", "field@nrpt.com", ""),

        # Branch 3: Softscape / Planting
        ("", "Softscape / Planting", "milestone", 20, "in_progress", "2025-02-15", "2025-04-30", "field@nrpt.com", "supervisor@nrpt.com"),
        ("Softscape / Planting", "Tree plantation", "sub_milestone", 60, "in_progress", "2025-02-15", "2025-04-10", "field@nrpt.com", ""),
        ("Softscape / Planting -> Tree plantation", "Pit digging", "task", 30, "completed", "2025-02-15", "2025-02-25", "field@nrpt.com", ""),
        ("Softscape / Planting -> Tree plantation", "Planting", "task", 50, "pending_approval", "2025-02-26", "2025-03-20", "field@nrpt.com", ""),
        ("Softscape / Planting -> Tree plantation", "Staking & watering", "task", 20, "in_progress", "2025-03-21", "2025-04-10", "field@nrpt.com", ""),
        ("Softscape / Planting", "Lawn development", "sub_milestone", 40, "pending", "2025-03-15", "2025-04-30", "field@nrpt.com", ""),
        ("Softscape / Planting -> Lawn development", "Soil preparation", "task", 40, "pending", "2025-03-15", "2025-03-25", "field@nrpt.com", ""),
        ("Softscape / Planting -> Lawn development", "Lawn laying", "task", 60, "pending", "2025-03-26", "2025-04-30", "field@nrpt.com", ""),

        # Branch 4: Irrigation
        ("", "Irrigation", "milestone", 30, "in_progress", "2025-01-20", "2025-04-15", "field@nrpt.com", "supervisor@nrpt.com"),
        ("Irrigation", "Mainline & valves", "sub_milestone", 50, "in_progress", "2025-01-20", "2025-02-20", "field@nrpt.com", ""),
        ("Irrigation -> Mainline & valves", "Pipe trenching", "task", 50, "completed", "2025-01-20", "2025-01-30", "field@nrpt.com", ""),
        ("Irrigation -> Mainline & valves", "Pipe laying + valves", "task", 50, "in_progress", "2025-01-31", "2025-02-20", "field@nrpt.com", ""),
        ("Irrigation", "Drip installation", "sub_milestone", 50, "pending", "2025-02-21", "2025-04-15", "field@nrpt.com", ""),
        ("Irrigation -> Drip installation", "Drip line laying", "task", 70, "pending", "2025-02-21", "2025-03-25", "field@nrpt.com", ""),
        ("Irrigation -> Drip installation", "Testing & commissioning", "task", 30, "pending", "2025-03-26", "2025-04-15", "field@nrpt.com", ""),
    ]

    # Styling and colors
    fill_milestone = PatternFill("solid", fgColor="DFF2E1")     # light green
    fill_sub = PatternFill("solid", fgColor="FFF4CC")           # light yellow
    fill_task = PatternFill("solid", fgColor="DDEBFF")          # light blue
    align = Alignment(vertical="center", wrap_text=True)

    for row_idx, row in enumerate(sample, start=2):
        for col_idx, val in enumerate(row, start=1):
            c = ws_data.cell(row=row_idx, column=col_idx, value=val)
            c.alignment = align
        t = (row[2] or "").strip().lower()
        if t == "milestone":
            fill = fill_milestone
        elif t == "sub_milestone":
            fill = fill_sub
        else:
            fill = fill_task
        for col_idx in range(1, len(headers) + 1):
            ws_data.cell(row=row_idx, column=col_idx).fill = fill

    ws_data.freeze_panes = "A2"
    ws_data.auto_filter.ref = f"A1:I{len(sample) + 1}"
    # Column widths
    ws_data.column_dimensions["A"].width = 40  # Parent_Path
    ws_data.column_dimensions["B"].width = 32  # Name
    ws_data.column_dimensions["C"].width = 14  # Type
    ws_data.column_dimensions["D"].width = 10  # Weight
    ws_data.column_dimensions["E"].width = 16  # Status
    ws_data.column_dimensions["F"].width = 12  # Start
    ws_data.column_dimensions["G"].width = 12  # End
    ws_data.column_dimensions["H"].width = 22  # Primary
    ws_data.column_dimensions["I"].width = 22  # Secondary

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def parse_wbs_excel_and_load(session: Session, project_id: str, file_bytes: bytes) -> int:
    """Parse WBS_Data sheet and create WBS items. Returns count created. Parent path resolved in order."""
    wb = openpyxl_load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    if "WBS_Data" not in wb.sheetnames:
        raise ValueError("Sheet 'WBS_Data' not found. Use the downloaded template.")
    ws = wb["WBS_Data"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    users = session.exec(select(User)).all()
    by_email = {u.email.strip().lower(): u for u in users}
    by_name = {u.name.strip(): u for u in users}
    path_to_id: Dict[str, str] = {}  # "Parent -> Child" -> id
    created = 0
    for row in rows:
        if not row or len(row) < 2:
            continue
        parent_path = (row[0] or "").strip() if row[0] is not None else ""
        name = (row[1] or "").strip() if row[1] is not None else ""
        if not name:
            continue
        item_type_str = (row[2] or "task").strip().lower() if row[2] is not None else "task"
        try:
            item_type = WbsItemType(item_type_str)
        except ValueError:
            item_type = WbsItemType.task
        weight = float(row[3]) if row[3] is not None and str(row[3]).strip() else 0.0
        try:
            weight = float(weight)
        except (TypeError, ValueError):
            weight = 0.0
        status_str = (row[4] or "pending").strip().lower() if row[4] is not None else "pending"
        try:
            status = WbsStatus(status_str)
        except ValueError:
            status = WbsStatus.pending
        start_date = str(row[5]).strip() if row[5] is not None else ""
        end_date = str(row[6]).strip() if row[6] is not None else ""
        if start_date and len(start_date) > 10:
            start_date = start_date[:10]
        if end_date and len(end_date) > 10:
            end_date = end_date[:10]
        primary_owner_id = None
        if row[7] and str(row[7]).strip():
            s = str(row[7]).strip()
            primary_owner_id = by_email.get(s.lower()) or by_name.get(s)
            if primary_owner_id:
                primary_owner_id = primary_owner_id.id
        secondary_owner_id = None
        if row[8] and str(row[8]).strip():
            s = str(row[8]).strip()
            u = by_email.get(s.lower()) or by_name.get(s)
            if u:
                secondary_owner_id = u.id
        parent_id = path_to_id.get(parent_path) if parent_path else None
        item = WbsItem(
            project_id=project_id,
            parent_id=parent_id,
            name=name,
            item_type=item_type,
            weight=weight,
            status=status,
            start_date=start_date or None,
            end_date=end_date or None,
            primary_owner_id=primary_owner_id,
            secondary_owner_id=secondary_owner_id,
            created_at=_now(),
            updated_at=_now(),
        )
        session.add(item)
        session.flush()
        full_path = f"{parent_path} -> {name}" if parent_path else name
        path_to_id[full_path] = item.id
        created += 1
    return created


@app.get("/wbs/template")
def wbs_download_template(
    project_id: Optional[str] = None,
    user: User = Depends(require_roles(Role.architect, Role.project_owner, Role.supervisor, Role.field_manager)),
    session: Session = Depends(get_session),
):
    """Download WBS Excel template with instructions and sample data."""
    if not project_id:
        projects = session.exec(select(Project).order_by(Project.name.asc())).all()
        project_id = projects[0].id if projects else ""
    content = build_wbs_excel_template(session, project_id)
    filename = "WBS_Upload_Template.xlsx"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/wbs/upload")
def wbs_upload(
    project_id: str = Form(...),
    file: UploadFile = File(...),
    user: User = Depends(require_roles(Role.architect, Role.project_owner, Role.supervisor)),
    session: Session = Depends(get_session),
):
    """Upload filled WBS Excel; creates WBS items for the project."""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Please upload an Excel file (.xlsx).")
    raw = file.file.read()
    if len(raw) > 5 * 1024 * 1024:
        raise HTTPException(400, "File too large.")
    try:
        count = parse_wbs_excel_and_load(session, project_id, raw)
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        raise HTTPException(400, f"Invalid or unsupported Excel: {e!s}")
    session.commit()
    return RedirectResponse(f"/wbs?project_id={project_id}&uploaded={count}", status_code=303)


def _boq_excel_instructions() -> List[str]:
    return [
        "BOQ/BOM UPLOAD TEMPLATE - INSTRUCTIONS",
        "",
        "1. Fill in the 'BOQ_Data' sheet with your materials per activity (see column headers).",
        "2. WBS_Path must match exactly a task/milestone path from your project's WBS (e.g. 'Earthwork & Drainage -> Grading and levelling').",
        "   Leave WBS_Path empty to attach the line to 'Unassigned'.",
        "3. Save as .xlsx and use the BOQ page Upload button.",
        "",
        "COLUMNS:",
        "- WBS_Path: Full path of the WBS item (copy from WBS screen or template sample). Optional.",
        "- Material_Name: Name of material (required). Creates Material Master if missing.",
        "- Unit: e.g. pcs, m, kg, cum (default pcs).",
        "- Estimated_Qty: Planned quantity.",
        "- Unit_Price: Price per unit.",
        "- Actual_Qty: Delivered/used quantity (optional, can be 0).",
    ]


def build_boq_excel_template(session: Session, project_id: str) -> bytes:
    """Build Excel with Instructions and BOQ_Data sheet (headers + sample rows)."""
    wb = Workbook()
    ws_inst = wb.active
    ws_inst.title = "Instructions"
    for row_idx, line in enumerate(_boq_excel_instructions(), start=1):
        ws_inst.cell(row=row_idx, column=1, value=line)
    ws_inst.column_dimensions["A"].width = 80
    path_to_id = wbs_path_to_id_map(session, project_id)
    sample_paths = list(path_to_id.keys())[:5] if path_to_id else ["(Add WBS first, then download template again)"]
    ws_data = wb.create_sheet("BOQ_Data", 1)
    headers = ["WBS_Path", "Material_Name", "Unit", "Estimated_Qty", "Unit_Price", "Actual_Qty"]
    for col, h in enumerate(headers, start=1):
        c = ws_data.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F2A44")
        c.alignment = Alignment(horizontal="center", vertical="center")
    sample = [
        (sample_paths[0] if sample_paths else "", "Bricks", "pcs", 5000, 12.5, 0),
        (sample_paths[0] if sample_paths else "", "Cement", "bags", 200, 450, 0),
        ("", "Labour (skilled)", "days", 30, 800, 0),
    ]
    for row_idx, row in enumerate(sample, start=2):
        for col_idx, val in enumerate(row, start=1):
            ws_data.cell(row=row_idx, column=col_idx, value=val)
    ws_data.column_dimensions["A"].width = 45
    ws_data.column_dimensions["B"].width = 24
    for col in "CDEF":
        ws_data.column_dimensions[col].width = 14
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def parse_boq_excel_and_load(
    session: Session, project_id: str, file_bytes: bytes, user_role: Role
) -> int:
    """Parse BOQ_Data sheet; create MaterialMaster if needed and BoqItem. Returns count created."""
    wb = openpyxl_load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    if "BOQ_Data" not in wb.sheetnames:
        raise ValueError("Sheet 'BOQ_Data' not found. Use the downloaded template.")
    ws = wb["BOQ_Data"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    path_to_id = wbs_path_to_id_map(session, project_id)
    materials_by_name = {m.name.strip(): m for m in session.exec(select(MaterialMaster)).all()}
    created = 0
    needs_approval = user_role in (Role.supervisor, Role.field_manager)
    for row in rows:
        if not row or len(row) < 2:
            continue
        wbs_path = (row[0] or "").strip() if row[0] is not None else ""
        material_name = (row[1] or "").strip() if row[1] is not None else ""
        if not material_name:
            continue
        unit = (row[2] or "pcs").strip() if len(row) > 2 and row[2] is not None else "pcs"
        try:
            est_qty = float(row[3]) if len(row) > 3 and row[3] is not None else 0.0
        except (TypeError, ValueError):
            est_qty = 0.0
        try:
            unit_price = float(row[4]) if len(row) > 4 and row[4] is not None else 0.0
        except (TypeError, ValueError):
            unit_price = 0.0
        try:
            actual_qty = float(row[5]) if len(row) > 5 and row[5] is not None else 0.0
        except (TypeError, ValueError):
            actual_qty = 0.0
        wbs_item_id = path_to_id.get(wbs_path) if wbs_path else None
        if material_name not in materials_by_name:
            m = MaterialMaster(name=material_name, default_unit=unit or "pcs")
            session.add(m)
            session.flush()
            materials_by_name[material_name] = m
        item = BoqItem(
            project_id=project_id,
            wbs_item_id=wbs_item_id,
            material_name=material_name,
            unit=unit or "pcs",
            estimated_quantity=est_qty,
            unit_price=unit_price,
            actual_quantity=actual_qty,
            pending_approval=needs_approval,
            created_at=_now(),
            updated_at=_now(),
        )
        session.add(item)
        created += 1
    return created


def compute_boq_rollup_by_wbs(session: Session, project_id: str) -> Dict[str, Dict[str, float]]:
    """Per wbs_item_id (and None for unassigned): estimated_cost, actual_cost, variance."""
    items = session.exec(select(BoqItem).where(BoqItem.project_id == project_id)).all()
    rollup: Dict[Optional[str], Dict[str, float]] = {}
    for i in items:
        key = i.wbs_item_id
        if key not in rollup:
            rollup[key] = {"estimated_cost": 0.0, "actual_cost": 0.0, "variance": 0.0}
        est = (i.estimated_quantity or 0.0) * (i.unit_price or 0.0)
        act = (i.actual_quantity or 0.0) * (i.unit_price or 0.0)
        rollup[key]["estimated_cost"] += est
        rollup[key]["actual_cost"] += act
        rollup[key]["variance"] += est - act
    return rollup


@app.get("/boq/template")
def boq_download_template(
    project_id: Optional[str] = None,
    user: User = Depends(require_roles(Role.architect, Role.project_owner, Role.supervisor, Role.field_manager)),
    session: Session = Depends(get_session),
):
    """Download BOQ/BOM Excel template with instructions and sample data."""
    if not project_id:
        projects = session.exec(select(Project).order_by(Project.name.asc())).all()
        project_id = projects[0].id if projects else ""
    content = build_boq_excel_template(session, project_id)
    filename = "BOQ_Upload_Template.xlsx"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/boq/upload")
def boq_upload(
    project_id: str = Form(...),
    file: UploadFile = File(...),
    user: User = Depends(require_roles(Role.architect, Role.project_owner, Role.supervisor)),
    session: Session = Depends(get_session),
):
    """Upload filled BOQ Excel; creates BOQ lines (and materials if needed) for the project."""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Please upload an Excel file (.xlsx).")
    raw = file.file.read()
    if len(raw) > 5 * 1024 * 1024:
        raise HTTPException(400, "File too large.")
    try:
        count = parse_boq_excel_and_load(session, project_id, raw, user.role)
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        raise HTTPException(400, f"Invalid or unsupported Excel: {e!s}")
    session.commit()
    return RedirectResponse(f"/boq?project_id={project_id}&uploaded={count}", status_code=303)


@app.get("/boq", response_class=HTMLResponse)
def boq_page(
    request: Request,
    project_id: Optional[str] = None,
    wbs_filter: Optional[str] = None,
    material_filter: Optional[str] = None,
    user: User = Depends(require_roles(Role.architect, Role.project_owner, Role.supervisor, Role.field_manager)),
    session: Session = Depends(get_session),
):
    projects = session.exec(select(Project).order_by(Project.name.asc())).all()
    if not project_id and projects:
        project_id = projects[0].id
    items = session.exec(select(BoqItem).where(BoqItem.project_id == (project_id or ""))).all() if project_id else []
    wbs_items = session.exec(select(WbsItem).where(WbsItem.project_id == (project_id or ""))).all() if project_id else []
    wbs_tasks = [w for w in wbs_items if w.item_type == WbsItemType.task]
    wbs_by_id = {w.id: w for w in wbs_items}
    costs = compute_project_costs(session, project_id) if project_id else {"estimated_cost": 0.0, "actual_cost": 0.0, "variance": 0.0}
    rollup_by_wbs = compute_boq_rollup_by_wbs(session, project_id) if project_id else {}
    project = session.exec(select(Project).where(Project.id == project_id)).first() if project_id else None
    budget = float(project.budget or 0.0) if project else 0.0
    grouped_boq: List[Dict[str, Any]] = []
    by_wbs: Dict[Optional[str], List[BoqItem]] = {}
    for i in items:
        by_wbs.setdefault(i.wbs_item_id, []).append(i)
    for wbs_id, group_items in by_wbs.items():
        if material_filter:
            group_items = [x for x in group_items if (x.material_name or "").strip() == material_filter]
        if not group_items:
            continue
        wbs_display = wbs_display_path(wbs_id, wbs_by_id)
        wbs_name = (wbs_by_id[wbs_id].name if wbs_id and wbs_id in wbs_by_id else "Unassigned") if wbs_id else "Unassigned"
        if material_filter:
            est = sum((x.estimated_quantity or 0.0) * (x.unit_price or 0.0) for x in group_items)
            act = sum((x.actual_quantity or 0.0) * (x.unit_price or 0.0) for x in group_items)
            sub = {"estimated_cost": est, "actual_cost": act, "variance": est - act}
        else:
            sub = rollup_by_wbs.get(wbs_id, {"estimated_cost": 0.0, "actual_cost": 0.0, "variance": 0.0})
        grouped_boq.append({
            "wbs_id": wbs_id, "wbs_name": wbs_name, "wbs_display": wbs_display,
            "line_items": group_items, "subtotal": sub, "bom_count": len(group_items),
        })
    grouped_boq.sort(key=lambda x: (x["wbs_name"] == "Unassigned", x["wbs_name"]))
    if wbs_filter:
        grouped_boq = [g for g in grouped_boq if g["wbs_id"] == wbs_filter]
    wbs_dropdown_options = build_wbs_dropdown_options(wbs_items, wbs_by_id)
    materials = session.exec(select(MaterialMaster).order_by(MaterialMaster.name.asc())).all()
    materials_approved = [m for m in materials if not m.pending_approval]
    material_filter_options = sorted({(i.material_name or "").strip() for i in items if (i.material_name or "").strip()})
    ctx = ui_context(session, user)
    ctx.update(
        {
            "request": request,
            "projects": projects,
            "project_id": project_id,
            "wbs_filter": wbs_filter,
            "material_filter": material_filter,
            "material_filter_options": material_filter_options,
            "items": items,
            "tasks": wbs_tasks,
            "wbs_items": wbs_items,
            "wbs_by_id": wbs_by_id,
            "costs": costs,
            "rollup_by_wbs": rollup_by_wbs,
            "project_budget": budget,
            "grouped_boq": grouped_boq,
            "wbs_dropdown_options": wbs_dropdown_options,
            "materials": materials,
            "materials_approved": materials_approved,
            "mask_prices": user.role == Role.field_manager,
        }
    )
    return templates.TemplateResponse("boq.html", ctx)


@app.post("/materials/create")
def materials_create(
    name: str = Form(""),  # optional so we can redirect back with error instead of 422
    default_unit: str = Form("pcs"),
    redirect_to: str = Form("boq"),
    project_id: str = Form(""),
    user: User = Depends(require_roles(Role.architect, Role.project_owner, Role.supervisor, Role.field_manager)),
    session: Session = Depends(get_session),
):
    name = name.strip()
    if not name:
        msg = "Material name is required."
        if redirect_to == "boq" and project_id:
            return RedirectResponse(f"/boq?project_id={project_id}&error={quote(msg)}", status_code=303)
        return RedirectResponse(f"/boq?error={quote(msg)}", status_code=303)
    existing = session.exec(select(MaterialMaster).where(MaterialMaster.name == name)).first()
    if existing:
        if redirect_to == "boq" and project_id:
            return RedirectResponse(f"/boq?project_id={project_id}&error={quote('Material already exists.')}", status_code=303)
        return RedirectResponse(f"/boq?error={quote('Material already exists.')}", status_code=303)
    # All new materials go through approval (show in Approvals until approved)
    m = MaterialMaster(name=name, default_unit=(default_unit or "pcs").strip(), pending_approval=True)
    session.add(m)
    session.commit()
    if redirect_to == "boq" and project_id:
        return RedirectResponse(f"/boq?project_id={project_id}", status_code=303)
    return RedirectResponse("/boq", status_code=303)


@app.get("/materials/{material_id}/edit", response_class=HTMLResponse)
def material_edit_page(
    material_id: str,
    request: Request,
    project_id: str = "",
    user: User = Depends(require_roles(Role.architect, Role.project_owner, Role.supervisor, Role.field_manager)),
    session: Session = Depends(get_session),
):
    m = session.exec(select(MaterialMaster).where(MaterialMaster.id == material_id)).first()
    if not m:
        raise HTTPException(404, "Material not found")
    ctx = ui_context(session, user)
    ctx.update({"request": request, "material": m, "project_id": project_id})
    return templates.TemplateResponse("material_edit.html", ctx)


@app.post("/materials/{material_id}/update")
def material_update(
    material_id: str,
    name: str = Form(""),
    default_unit: str = Form("pcs"),
    project_id: str = Form(""),
    user: User = Depends(require_roles(Role.architect, Role.project_owner, Role.supervisor, Role.field_manager)),
    session: Session = Depends(get_session),
):
    m = session.exec(select(MaterialMaster).where(MaterialMaster.id == material_id)).first()
    if not m:
        raise HTTPException(404, "Material not found")
    name = name.strip()
    if not name:
        msg = "Material name is required."
        return RedirectResponse(f"/materials/{material_id}/edit?project_id={project_id}&error={quote(msg)}", status_code=303)
    existing = session.exec(select(MaterialMaster).where(MaterialMaster.name == name).where(MaterialMaster.id != material_id)).first()
    if existing:
        return RedirectResponse(f"/materials/{material_id}/edit?project_id={project_id}&error={quote('Name already in use.')}", status_code=303)
    m.name = name
    m.default_unit = (default_unit or "pcs").strip()
    # All material edits go through approval
    m.pending_approval = True
    session.add(m)
    session.commit()
    if project_id:
        return RedirectResponse(f"/boq?project_id={project_id}", status_code=303)
    return RedirectResponse("/boq", status_code=303)


@app.post("/materials/{material_id}/delete")
def material_delete(
    material_id: str,
    project_id: str = Form(""),
    user: User = Depends(require_roles(Role.architect, Role.project_owner)),
    session: Session = Depends(get_session),
):
    m = session.exec(select(MaterialMaster).where(MaterialMaster.id == material_id)).first()
    if not m:
        raise HTTPException(404, "Material not found")
    session.delete(m)
    session.commit()
    if project_id:
        return RedirectResponse(f"/boq?project_id={project_id}", status_code=303)
    return RedirectResponse("/boq", status_code=303)


@app.post("/boq/create")
def boq_create(
    project_id: str = Form(...),
    wbs_item_id: str = Form(""),
    material_name: str = Form(...),
    unit: str = Form("pcs"),
    estimated_quantity: float = Form(0.0),
    unit_price: float = Form(0.0),
    actual_quantity: float = Form(0.0),
    user: User = Depends(require_roles(Role.architect, Role.project_owner, Role.supervisor, Role.field_manager)),
    session: Session = Depends(get_session),
):
    # Basic guards: no negative quantities or prices
    est = float(estimated_quantity or 0.0)
    price = float(unit_price or 0.0)
    act = float(actual_quantity or 0.0)
    if est < 0 or price < 0 or act < 0:
        msg = "Quantities and unit price cannot be negative."
        return RedirectResponse(f"/boq?project_id={project_id}&error={quote(msg)}", status_code=303)
    needs_approval = user.role in (Role.supervisor, Role.field_manager)
    item = BoqItem(
        project_id=project_id,
        wbs_item_id=wbs_item_id or None,
        material_name=material_name.strip(),
        unit=unit.strip(),
        estimated_quantity=est,
        unit_price=price,
        actual_quantity=act,
        pending_approval=needs_approval,
        created_at=_now(),
        updated_at=_now(),
    )
    session.add(item)
    session.commit()
    return RedirectResponse(f"/boq?project_id={project_id}", status_code=303)


@app.post("/boq/{item_id}/actual")
def boq_update_actual(
    item_id: str,
    project_id: str = Form(...),
    actual_quantity: float = Form(0.0),
    user: User = Depends(require_roles(Role.architect, Role.project_owner, Role.supervisor, Role.field_manager)),
    session: Session = Depends(get_session),
):
    item = session.exec(select(BoqItem).where(BoqItem.id == item_id)).first()
    if not item:
        raise HTTPException(404, "BOQ item not found")
    act = float(actual_quantity or 0.0)
    if act < 0:
        msg = "Actual quantity cannot be negative."
        return RedirectResponse(f"/boq?project_id={project_id}&error={quote(msg)}", status_code=303)
    item.actual_quantity = act
    item.updated_at = _now()
    # Supervisor / Field actual qty changes go to approval queue
    if user.role in (Role.field_manager, Role.supervisor):
        item.pending_approval = True
    session.add(item)
    session.commit()
    return RedirectResponse(f"/boq?project_id={project_id}", status_code=303)


@app.get("/boq/{item_id}/edit", response_class=HTMLResponse)
def boq_edit_page(
    item_id: str,
    request: Request,
    project_id: Optional[str] = None,
    wbs_filter: Optional[str] = None,
    user: User = Depends(require_roles(Role.architect, Role.project_owner, Role.supervisor, Role.field_manager)),
    session: Session = Depends(get_session),
):
    item = session.exec(select(BoqItem).where(BoqItem.id == item_id)).first()
    if not item:
        raise HTTPException(404, "BOQ item not found")
    project_id = project_id or item.project_id
    wbs_items = session.exec(select(WbsItem).where(WbsItem.project_id == item.project_id)).all()
    wbs_by_id = {w.id: w for w in wbs_items}
    wbs_dropdown_options = build_wbs_dropdown_options(wbs_items, wbs_by_id)
    materials = session.exec(select(MaterialMaster).order_by(MaterialMaster.name.asc())).all()
    ctx = ui_context(session, user)
    ctx.update({
        "request": request, "item": item, "project_id": project_id, "wbs_filter": wbs_filter or "",
        "wbs_dropdown_options": wbs_dropdown_options, "materials": materials,
    })
    return templates.TemplateResponse("boq_edit.html", ctx)


@app.post("/boq/{item_id}/update")
def boq_update(
    item_id: str,
    project_id: str = Form(...),
    wbs_filter: str = Form(""),
    wbs_item_id: str = Form(""),
    material_name: str = Form(...),
    unit: str = Form("pcs"),
    estimated_quantity: float = Form(0.0),
    unit_price: float = Form(0.0),
    actual_quantity: float = Form(0.0),
    user: User = Depends(require_roles(Role.architect, Role.project_owner, Role.supervisor, Role.field_manager)),
    session: Session = Depends(get_session),
):
    item = session.exec(select(BoqItem).where(BoqItem.id == item_id)).first()
    if not item:
        raise HTTPException(404, "BOQ item not found")
    est = float(estimated_quantity or 0.0)
    price = float(unit_price or 0.0)
    act = float(actual_quantity or 0.0)
    if est < 0 or price < 0 or act < 0:
        msg = "Quantities and unit price cannot be negative."
        return RedirectResponse(f"/boq/{item_id}/edit?project_id={project_id}&error={quote(msg)}", status_code=303)
    item.wbs_item_id = (wbs_item_id or "").strip() or None
    item.material_name = material_name.strip()
    item.unit = (unit or "pcs").strip()
    item.estimated_quantity = est
    item.unit_price = price
    item.actual_quantity = act
    item.updated_at = _now()
    if user.role in (Role.supervisor, Role.field_manager):
        item.pending_approval = True
    session.add(item)
    session.commit()
    url = f"/boq?project_id={project_id}"
    if (wbs_filter or "").strip():
        url += f"&wbs_filter={quote((wbs_filter or '').strip())}"
    return RedirectResponse(url, status_code=303)


@app.post("/boq/{item_id}/delete")
def boq_delete(
    item_id: str,
    project_id: str = Form(...),
    wbs_filter: str = Form(""),
    user: User = Depends(require_roles(Role.architect, Role.project_owner)),
    session: Session = Depends(get_session),
):
    item = session.exec(select(BoqItem).where(BoqItem.id == item_id)).first()
    if not item:
        raise HTTPException(404, "BOQ item not found")
    pid = item.project_id
    session.delete(item)
    session.commit()
    url = f"/boq?project_id={project_id or pid}"
    if (wbs_filter or "").strip():
        url += f"&wbs_filter={quote((wbs_filter or '').strip())}"
    return RedirectResponse(url, status_code=303)


@app.get("/defects", response_class=HTMLResponse)
def defects_page(
    request: Request,
    project_id: Optional[str] = None,
    wbs_item_id: Optional[str] = None,
    location: Optional[str] = None,
    wbs_filter: Optional[str] = None,
    severity_filter: Optional[str] = None,
    status_filter: Optional[str] = None,
    assigned_filter: Optional[str] = None,
    user: User = Depends(require_roles(Role.architect, Role.project_owner, Role.supervisor, Role.field_manager)),
    session: Session = Depends(get_session),
):
    projects = session.exec(select(Project).order_by(Project.name.asc())).all()
    if not project_id and projects:
        project_id = projects[0].id
    defects = session.exec(select(Defect).where(Defect.project_id == (project_id or "")).order_by(Defect.created_at.desc())).all() if project_id else []
    if wbs_filter:
        defects = [d for d in defects if getattr(d, "wbs_item_id", None) == wbs_filter]
    if severity_filter:
        defects = [d for d in defects if getattr(d.severity, "value", str(d.severity)) == severity_filter]
    if status_filter:
        defects = [d for d in defects if getattr(d.status, "value", str(d.status)) == status_filter]
    if assigned_filter:
        defects = [d for d in defects if (d.assigned_to_user_id or "") == assigned_filter]
    wbs_items = session.exec(select(WbsItem).where(WbsItem.project_id == (project_id or ""))).all() if project_id else []
    users = session.exec(select(User).order_by(User.name.asc())).all()
    users_by_id = {u.id: u for u in users}
    wbs_by_id = {w.id: w for w in wbs_items}
    defects_with_attachments = []
    for d in defects:
        try:
            atts = session.exec(select(DefectAttachment).where(DefectAttachment.defect_id == d.id)).all()
        except Exception:
            atts = []
        wid = getattr(d, "wbs_item_id", None)
        wbs_name = (wbs_by_id.get(wid).name if wid and wid in wbs_by_id else None)
        sev_val = getattr(d.severity, "value", None) or str(d.severity).split(".")[-1] if d.severity else "medium"
        status_val = getattr(d.status, "value", None) or str(d.status).split(".")[-1] if d.status else "open"
        defects_with_attachments.append({
            "defect": d, "attachments": atts, "wbs_name": wbs_name, "attachment_count": len(atts),
            "severity_display": sev_val, "status_display": status_val,
        })
    # Pre-fill form when coming from WBS "Create defect" (validate wbs belongs to project)
    prefill: Optional[Dict[str, Any]] = None
    if wbs_item_id and (project_id or (projects and projects[0].id)):
        pid = project_id or (projects[0].id if projects else "")
        wbs_item = session.exec(select(WbsItem).where(WbsItem.id == wbs_item_id)).first()
        if wbs_item and wbs_item.project_id == pid:
            prefill = {
                "project_id": pid,
                "wbs_item_id": wbs_item_id,
                "wbs_name": wbs_item.name,
                "location": (location or "").strip() or wbs_item.name,
                "reported_by": user.name or "",
                "assigned_to_user_id": wbs_item.primary_owner_id or "",
            }
    wbs_dropdown_options = build_wbs_dropdown_options(wbs_items, wbs_by_id)
    ctx = ui_context(session, user)
    ctx.update({
        "request": request,
        "projects": projects,
        "project_id": project_id,
        "wbs_filter": wbs_filter,
        "severity_filter": severity_filter,
        "status_filter": status_filter,
        "assigned_filter": assigned_filter,
        "defects": defects_with_attachments,
        "wbs_items": wbs_items,
        "wbs_dropdown_options": wbs_dropdown_options,
        "users": users,
        "users_by_id": users_by_id,
        "prefill": prefill,
    })
    return templates.TemplateResponse("defects.html", ctx)


@app.post("/defects/create")
async def defects_create(
    request: Request,
    user: User = Depends(require_roles(Role.architect, Role.project_owner, Role.supervisor, Role.field_manager)),
    session: Session = Depends(get_session),
):
    try:
        form = await request.form()
        project_id = (form.get("project_id") or "").strip()
        if not project_id:
            return RedirectResponse("/defects?error=missing_project", status_code=303)
        location = (form.get("location") or "").strip()
        description = (form.get("description") or "").strip()
        if not location or not description:
            return RedirectResponse(f"/defects?project_id={project_id}&error=missing_fields", status_code=303)
        wbs_item_id = (form.get("wbs_item_id") or "").strip() or None
        reported_by = (form.get("reported_by") or "").strip()
        reporter_contact = (form.get("reporter_contact") or "").strip()
        assigned_to_user_id = (form.get("assigned_to_user_id") or "").strip() or None
        try:
            severity = DefectSeverity((form.get("severity") or "medium").strip().lower())
        except ValueError:
            severity = DefectSeverity.medium
        try:
            status = DefectStatus((form.get("status") or "open").strip().lower())
        except ValueError:
            status = DefectStatus.open
        r = session.exec(select(sqlfunc.coalesce(sqlfunc.max(Defect.display_number), 0)).where(Defect.project_id == project_id)).first()
        next_num = (r or 0) + 1
        d = Defect(
            project_id=project_id,
            wbs_item_id=wbs_item_id,
            display_number=next_num,
            location=location,
            description=description,
            severity=severity,
            status=status,
            reported_by=reported_by,
            reporter_contact=reporter_contact,
            assigned_to_user_id=assigned_to_user_id,
            created_at=_now(),
            updated_at=_now(),
        )
        session.add(d)
        session.commit()
        session.refresh(d)
        files_list = form.getlist("files") if hasattr(form, "getlist") else []
        for f in files_list or []:
            fn = getattr(f, "filename", None)
            ct = getattr(f, "content_type", None)
            if not fn or not ct:
                continue
            # Support both async and sync read (form uploads may differ by framework)
            read_fn = getattr(f, "read", None)
            if read_fn is None:
                continue
            try:
                if asyncio.iscoroutinefunction(read_fn):
                    raw = await read_fn()
                else:
                    raw = read_fn()
            except Exception:
                raw = b""
            if not raw or len(raw) > 10 * 1024 * 1024:  # 10 MB
                continue
            ft = "photo"
            if ct.startswith("video/"):
                ft = "video"
            elif ct.startswith("audio/"):
                ft = "audio"
            att = DefectAttachment(
                defect_id=d.id,
                file_type=DefectAttachmentType(ft),
                filename=fn or "file",
                content_type=ct or "application/octet-stream",
                content_base64=base64.b64encode(raw).decode("ascii"),
                phase="before",
                created_at=_now(),
            )
            session.add(att)
        if files_list:
            session.commit()
        return RedirectResponse(f"/defects?project_id={project_id}", status_code=303)
    except Exception as e:
        tb = traceback.format_exc()
        err_esc = str(e).replace("<", "&lt;").replace(">", "&gt;")
        tb_esc = tb.replace("<", "&lt;").replace(">", "&gt;")
        return HTMLResponse(
            f"<html><body style='font-family:sans-serif;padding:20px;'><h2>Defect create error</h2><pre>{err_esc}</pre><pre>{tb_esc}</pre><p><a href='/defects'>Back to Defects</a></p></body></html>",
            status_code=500,
        )


@app.get("/defects/{defect_id}/edit", response_class=HTMLResponse)
def defect_edit_page(
    defect_id: str,
    request: Request,
    project_id: Optional[str] = None,
    wbs_filter: str = "",
    severity_filter: str = "",
    status_filter: str = "",
    assigned_filter: str = "",
    user: User = Depends(require_roles(Role.architect, Role.project_owner)),
    session: Session = Depends(get_session),
):
    d = session.exec(select(Defect).where(Defect.id == defect_id)).first()
    if not d:
        raise HTTPException(404, "Defect not found")
    project_id = project_id or d.project_id
    wbs_items = session.exec(select(WbsItem).where(WbsItem.project_id == d.project_id)).all()
    wbs_by_id = {w.id: w for w in wbs_items}
    wbs_dropdown_options = build_wbs_dropdown_options(wbs_items, wbs_by_id)
    users = session.exec(select(User).order_by(User.name.asc())).all()
    ctx = ui_context(session, user)
    ctx.update({
        "request": request, "defect": d, "project_id": project_id,
        "wbs_dropdown_options": wbs_dropdown_options, "users": users,
        "wbs_filter": wbs_filter or "", "severity_filter": severity_filter or "",
        "status_filter": status_filter or "", "assigned_filter": assigned_filter or "",
    })
    return templates.TemplateResponse("defect_edit.html", ctx)


@app.post("/defects/{defect_id}/update")
def defect_update(
    defect_id: str,
    project_id: str = Form(...),
    wbs_item_id: str = Form(""),
    location: str = Form(...),
    description: str = Form(...),
    severity: DefectSeverity = Form(DefectSeverity.medium),
    status: DefectStatus = Form(DefectStatus.open),
    reported_by: str = Form(""),
    reporter_contact: str = Form(""),
    assigned_to_user_id: str = Form(""),
    wbs_filter: str = Form(""),
    severity_filter: str = Form(""),
    status_filter: str = Form(""),
    assigned_filter: str = Form(""),
    user: User = Depends(require_roles(Role.architect, Role.project_owner)),
    session: Session = Depends(get_session),
):
    d = session.exec(select(Defect).where(Defect.id == defect_id)).first()
    if not d:
        raise HTTPException(404, "Defect not found")
    d.wbs_item_id = (wbs_item_id or "").strip() or None
    d.location = location.strip()
    d.description = description.strip()
    d.severity = severity
    d.status = status
    d.reported_by = (reported_by or "").strip()
    d.reporter_contact = (reporter_contact or "").strip()
    d.assigned_to_user_id = (assigned_to_user_id or "").strip() or None
    d.updated_at = _now()
    session.add(d)
    session.commit()
    return RedirectResponse(
        "/defects" + _defect_list_query(project_id, wbs_filter, severity_filter, status_filter, assigned_filter),
        status_code=303,
    )


@app.post("/defects/{defect_id}/delete")
def defect_delete(
    defect_id: str,
    project_id: str = Form(...),
    wbs_filter: str = Form(""),
    severity_filter: str = Form(""),
    status_filter: str = Form(""),
    assigned_filter: str = Form(""),
    user: User = Depends(require_roles(Role.architect, Role.project_owner)),
    session: Session = Depends(get_session),
):
    d = session.exec(select(Defect).where(Defect.id == defect_id)).first()
    if not d:
        raise HTTPException(404, "Defect not found")
    pid = d.project_id
    for att in session.exec(select(DefectAttachment).where(DefectAttachment.defect_id == defect_id)).all():
        session.delete(att)
    session.delete(d)
    session.commit()
    return RedirectResponse(
        "/defects" + _defect_list_query(project_id or pid, wbs_filter, severity_filter, status_filter, assigned_filter),
        status_code=303,
    )


@app.post("/defects/{defect_id}/status")
def defect_update_status(
    defect_id: str,
    project_id: str = Form(...),
    status: DefectStatus = Form(...),
    wbs_filter: str = Form(""),
    severity_filter: str = Form(""),
    status_filter: str = Form(""),
    assigned_filter: str = Form(""),
    user: User = Depends(require_roles(Role.architect, Role.project_owner, Role.supervisor, Role.field_manager)),
    session: Session = Depends(get_session),
):
    d = session.exec(select(Defect).where(Defect.id == defect_id)).first()
    if not d:
        raise HTTPException(404, "Defect not found")
    # Resolved goes to approval queue so supervisor can approve (before/after comparison)
    if status == DefectStatus.resolved:
        d.status = DefectStatus.pending_approval
    else:
        d.status = status
    d.updated_at = _now()
    session.add(d)
    session.commit()
    return RedirectResponse(
        "/defects" + _defect_list_query(project_id, wbs_filter, severity_filter, status_filter, assigned_filter),
        status_code=303,
    )


@app.post("/defects/{defect_id}/attachments")
def defect_upload_attachments(
    defect_id: str,
    project_id: str = Form(...),
    files: List[UploadFile] = File([]),
    phase: str = Form("before"),  # "before" or "after" for before/after comparison at resolution
    wbs_filter: str = Form(""),
    severity_filter: str = Form(""),
    status_filter: str = Form(""),
    assigned_filter: str = Form(""),
    user: User = Depends(require_roles(Role.architect, Role.project_owner, Role.supervisor, Role.field_manager)),
    session: Session = Depends(get_session),
):
    d = session.exec(select(Defect).where(Defect.id == defect_id)).first()
    if not d:
        raise HTTPException(404, "Defect not found")
    phase_val = "after" if (phase and phase.strip().lower() == "after") else "before"
    for f in files or []:
        if not f.filename or not f.content_type:
            continue
        raw = f.file.read()
        if len(raw) > 10 * 1024 * 1024:  # 10 MB limit
            continue
        ft = "photo"
        if f.content_type.startswith("video/"):
            ft = "video"
        elif f.content_type.startswith("audio/"):
            ft = "audio"
        att = DefectAttachment(
            defect_id=defect_id,
            file_type=DefectAttachmentType(ft),
            filename=f.filename or "file",
            content_type=f.content_type or "application/octet-stream",
            content_base64=base64.b64encode(raw).decode("ascii"),
            phase=phase_val,
            created_at=_now(),
        )
        session.add(att)
    session.commit()
    return RedirectResponse(
        "/defects" + _defect_list_query(project_id, wbs_filter, severity_filter, status_filter, assigned_filter),
        status_code=303,
    )


@app.get("/defects/attachments/{attachment_id}")
def defect_serve_attachment(
    attachment_id: str,
    session: Session = Depends(get_session),
):
    att = session.exec(select(DefectAttachment).where(DefectAttachment.id == attachment_id)).first()
    if not att or not att.content_base64:
        raise HTTPException(404, "Not found")
    try:
        body = base64.b64decode(att.content_base64)
    except Exception:
        raise HTTPException(404, "Invalid")
    return RawResponse(content=body, media_type=att.content_type or "application/octet-stream")


@app.get("/report-defect", response_class=HTMLResponse)
def report_defect_public(request: Request, session: Session = Depends(get_session)):
    projects = session.exec(select(Project).order_by(Project.name.asc())).all()
    return templates.TemplateResponse("report_defect.html", {"request": request, "projects": projects})


@app.post("/report-defect")
async def report_defect_public_submit(
    request: Request,
    session: Session = Depends(get_session),
):
    form = await request.form()
    project_id = (form.get("project_id") or "").strip()
    location = (form.get("location") or "").strip()
    description = (form.get("description") or "").strip()
    reported_by = (form.get("reported_by") or "").strip()
    reporter_contact = (form.get("reporter_contact") or "").strip()
    if not project_id or not location or not description:
        return RedirectResponse("/report-defect?error=missing", status_code=303)
    try:
        severity = DefectSeverity((form.get("severity") or "medium").strip().lower())
    except ValueError:
        severity = DefectSeverity.medium
    r = session.exec(select(sqlfunc.coalesce(sqlfunc.max(Defect.display_number), 0)).where(Defect.project_id == project_id)).first()
    next_num = (r or 0) + 1
    d = Defect(
        project_id=project_id,
        display_number=next_num,
        location=location,
        description=description,
        severity=severity,
        status=DefectStatus.open,
        reported_by=reported_by,
        reporter_contact=reporter_contact,
        created_at=_now(),
        updated_at=_now(),
    )
    session.add(d)
    session.commit()
    session.refresh(d)
    files_list = form.getlist("files") if hasattr(form, "getlist") else []
    for f in files_list or []:
        fn = getattr(f, "filename", None)
        ct = getattr(f, "content_type", None)
        if not fn or not ct:
            continue
        raw = await f.read()
        if len(raw) > 10 * 1024 * 1024:
            continue
        ft = "photo"
        if ct.startswith("video/"):
            ft = "video"
        elif ct.startswith("audio/"):
            ft = "audio"
        att = DefectAttachment(
            defect_id=d.id,
            file_type=DefectAttachmentType(ft),
            filename=fn or "file",
            content_type=ct or "application/octet-stream",
            content_base64=base64.b64encode(raw).decode("ascii"),
            phase="before",
            created_at=_now(),
        )
        session.add(att)
    if files_list:
        session.commit()
    return RedirectResponse("/report-defect?ok=1", status_code=303)


@app.get("/approvals", response_class=HTMLResponse)
def approvals_page(
    request: Request,
    user: User = Depends(require_roles(Role.architect, Role.project_owner, Role.supervisor)),
    session: Session = Depends(get_session),
):
    pending_wbs = session.exec(select(WbsItem).where(WbsItem.status == WbsStatus.pending_approval).order_by(WbsItem.updated_at.desc())).all()
    pending_defects = session.exec(select(Defect).where(Defect.status == DefectStatus.pending_approval).order_by(Defect.updated_at.desc())).all()
    pending_boq = session.exec(select(BoqItem).where(BoqItem.pending_approval == True).order_by(BoqItem.updated_at.desc())).all()
    pending_materials = session.exec(select(MaterialMaster).where(MaterialMaster.pending_approval == True).order_by(MaterialMaster.name.asc())).all()
    projects = {p.id: p for p in session.exec(select(Project)).all()}
    wbs_by_id = {}
    for w in session.exec(select(WbsItem)).all():
        wbs_by_id[w.id] = w
    ctx = ui_context(session, user)
    ctx.update({
        "request": request,
        "pending_wbs": pending_wbs,
        "pending_defects": pending_defects,
        "pending_boq": pending_boq,
        "pending_materials": pending_materials,
        "projects": projects,
        "wbs_by_id": wbs_by_id,
    })
    return templates.TemplateResponse("approvals.html", ctx)


@app.post("/approvals/wbs/{item_id}")
def approvals_wbs_action(
    item_id: str,
    action: str = Form(...),  # approve or reject
    reason: str = Form(""),
    user: User = Depends(require_roles(Role.architect, Role.project_owner, Role.supervisor)),
    session: Session = Depends(get_session),
):
    item = session.exec(select(WbsItem).where(WbsItem.id == item_id)).first()
    if not item:
        raise HTTPException(404, "WBS item not found")
    if action == "approve":
        item.status = WbsStatus.completed
    elif action == "reject":
        if not reason.strip():
            raise HTTPException(400, "Rejection reason required")
        item.status = WbsStatus.rejected
        item.name = f"{item.name} (Rejected: {reason.strip()})"
    else:
        raise HTTPException(400, "Invalid action")
    item.updated_at = _now()
    session.add(item)
    session.commit()
    return RedirectResponse("/approvals", status_code=303)


@app.post("/approvals/defect/{defect_id}")
def approvals_defect_action(
    defect_id: str,
    action: str = Form(...),  # approve or reject
    reason: str = Form(""),
    user: User = Depends(require_roles(Role.architect, Role.project_owner, Role.supervisor)),
    session: Session = Depends(get_session),
):
    d = session.exec(select(Defect).where(Defect.id == defect_id)).first()
    if not d:
        raise HTTPException(404, "Defect not found")
    if action == "approve":
        d.status = DefectStatus.approved
    elif action == "reject":
        if not reason.strip():
            raise HTTPException(400, "Rejection reason required")
        d.status = DefectStatus.reopened
        d.description = f"{d.description} [Rejected: {reason.strip()}]"
    else:
        raise HTTPException(400, "Invalid action")
    d.updated_at = _now()
    session.add(d)
    session.commit()
    return RedirectResponse("/approvals", status_code=303)


@app.post("/approvals/boq/{item_id}")
def approvals_boq_action(
    item_id: str,
    action: str = Form(...),  # approve or reject
    user: User = Depends(require_roles(Role.architect, Role.project_owner, Role.supervisor)),
    session: Session = Depends(get_session),
):
    item = session.exec(select(BoqItem).where(BoqItem.id == item_id)).first()
    if not item:
        raise HTTPException(404, "BOQ item not found")
    if action == "approve":
        item.pending_approval = False
    elif action == "reject":
        item.pending_approval = False
        item.actual_quantity = item.estimated_quantity or 0.0  # revert to estimated
    else:
        raise HTTPException(400, "Invalid action")
    item.updated_at = _now()
    session.add(item)
    session.commit()
    return RedirectResponse("/approvals", status_code=303)


@app.post("/approvals/material/{material_id}")
def approvals_material_action(
    material_id: str,
    action: str = Form(...),  # approve or reject
    user: User = Depends(require_roles(Role.architect, Role.project_owner, Role.supervisor)),
    session: Session = Depends(get_session),
):
    m = session.exec(select(MaterialMaster).where(MaterialMaster.id == material_id)).first()
    if not m:
        raise HTTPException(404, "Material not found")
    if action == "approve":
        m.pending_approval = False
    elif action == "reject":
        m.pending_approval = False  # clear from queue; material stays
    else:
        raise HTTPException(400, "Invalid action")
    session.add(m)
    session.commit()
    return RedirectResponse("/approvals", status_code=303)


@app.get("/users", response_class=HTMLResponse)
def users_page(
    request: Request,
    user: User = Depends(require_roles(Role.architect, Role.project_owner)),
    session: Session = Depends(get_session),
):
    users = session.exec(select(User).order_by(User.created_at.desc())).all()
    projects = session.exec(select(Project).order_by(Project.name.asc())).all()
    projects_by_id = {pr.id: pr for pr in projects}
    user_projects_map: Dict[str, List[str]] = {}
    user_locations_map: Dict[str, List[str]] = {}
    for u in users:
        ups = session.exec(select(UserProject).where(UserProject.user_id == u.id)).all()
        user_projects_map[u.id] = [projects_by_id[up.project_id].name for up in ups if up.project_id in projects_by_id]
        uls = session.exec(select(UserLocation).where(UserLocation.user_id == u.id)).all()
        user_locations_map[u.id] = [ul.location for ul in uls]
    perms = session.exec(select(RolePermission)).all()
    perm_by_role_resource: Dict[Tuple[str, str], RolePermission] = {}
    for p in perms:
        perm_by_role_resource[(p.role, p.resource)] = p
    ctx = ui_context(session, user)
    ctx.update({
        "request": request,
        "users": users,
        "projects": projects,
        "user_projects_map": user_projects_map,
        "user_locations_map": user_locations_map,
        "perm_by_role_resource": perm_by_role_resource,
        "roles": [r for r in Role],
        "resources": [r for r in PermissionResource],
    })
    return templates.TemplateResponse("users.html", ctx)


@app.post("/users/create")
def users_create(
    request: Request,
    name: str = Form(...),
    email: str = Form(...),
    password: str = Form(""),
    role: str = Form(...),
    phone: str = Form(""),
    whatsapp_phone: str = Form(""),
    address: str = Form(""),
    locations: str = Form(""),
    user: User = Depends(require_roles(Role.architect, Role.project_owner)),
    session: Session = Depends(get_session),
):
    form_data = request.form()
    project_ids = list(form_data.getlist("project_ids")) if form_data else []
    email_clean = email.strip().lower()
    if not email_clean:
        return RedirectResponse("/users?error=" + quote("Email required"), status_code=303)
    existing = session.exec(select(User).where(User.email == email_clean)).first()
    if existing:
        return RedirectResponse("/users?error=" + quote("Email already in use"), status_code=303)
    role_enum = getattr(Role, role, None) or Role.field_manager
    password_hash = hash_password(password) if password else hash_password("password")
    new_user = User(
        name=name.strip(),
        email=email_clean,
        role=role_enum,
        password_hash=password_hash,
        phone=(phone or "").strip(),
        whatsapp_phone=(whatsapp_phone or "").strip(),
        address=(address or "").strip(),
    )
    session.add(new_user)
    session.commit()
    session.refresh(new_user)
    for pid in project_ids:
        if pid and str(pid).strip():
            session.add(UserProject(user_id=new_user.id, project_id=str(pid).strip()))
    for loc in (locations or "").split("\n"):
        loc = loc.strip()
        if loc:
            session.add(UserLocation(user_id=new_user.id, location=loc))
    session.commit()
    return RedirectResponse("/users?created=1", status_code=303)


@app.get("/users/{user_id}/edit", response_class=HTMLResponse)
def user_edit_page(
    request: Request,
    user_id: str,
    user: User = Depends(require_roles(Role.architect, Role.project_owner)),
    session: Session = Depends(get_session),
):
    edit_user = session.exec(select(User).where(User.id == user_id)).first()
    if not edit_user:
        raise HTTPException(404, "User not found")
    projects = session.exec(select(Project).order_by(Project.name.asc())).all()
    user_projects = [up.project_id for up in session.exec(select(UserProject).where(UserProject.user_id == user_id)).all()]
    user_locations = [ul.location for ul in session.exec(select(UserLocation).where(UserLocation.user_id == user_id)).all()]
    ctx = ui_context(session, user)
    ctx.update({
        "request": request,
        "edit_user": edit_user,
        "projects": projects,
        "user_projects": user_projects,
        "user_locations": user_locations,
        "roles": [r for r in Role],
    })
    return templates.TemplateResponse("user_edit.html", ctx)


@app.post("/users/{user_id}/edit")
def user_edit(
    request: Request,
    user_id: str,
    name: str = Form(...),
    email: str = Form(...),
    role: str = Form(...),
    phone: str = Form(""),
    whatsapp_phone: str = Form(""),
    address: str = Form(""),
    locations: str = Form(""),
    change_password: str = Form(""),
    user: User = Depends(require_roles(Role.architect, Role.project_owner)),
    session: Session = Depends(get_session),
):
    form_data = request.form()
    project_ids = list(form_data.getlist("project_ids")) if form_data else []
    edit_user = session.exec(select(User).where(User.id == user_id)).first()
    if not edit_user:
        raise HTTPException(404, "User not found")
    email_clean = email.strip().lower()
    if not email_clean:
        return RedirectResponse(f"/users/{user_id}/edit?error=" + quote("Email required"), status_code=303)
    other = session.exec(select(User).where(User.email == email_clean)).first()
    if other and other.id != user_id:
        return RedirectResponse(f"/users/{user_id}/edit?error=" + quote("Email already in use"), status_code=303)
    edit_user.name = name.strip()
    edit_user.email = email_clean
    edit_user.role = getattr(Role, role, None) or edit_user.role
    edit_user.phone = (phone or "").strip()
    edit_user.whatsapp_phone = (whatsapp_phone or "").strip()
    edit_user.address = (address or "").strip()
    if change_password:
        edit_user.password_hash = hash_password(change_password)
    session.add(edit_user)
    session.commit()
    session.exec(delete(UserProject).where(UserProject.user_id == user_id))
    session.exec(delete(UserLocation).where(UserLocation.user_id == user_id))
    for pid in project_ids:
        if pid and str(pid).strip():
            session.add(UserProject(user_id=user_id, project_id=str(pid).strip()))
    for loc in (locations or "").split("\n"):
        loc = loc.strip()
        if loc:
            session.add(UserLocation(user_id=user_id, location=loc))
    session.commit()
    return RedirectResponse("/users?updated=1", status_code=303)


@app.post("/users/permissions")
async def users_permissions_save(
    request: Request,
    user: User = Depends(require_roles(Role.architect, Role.project_owner)),
    session: Session = Depends(get_session),
):
    form = await request.form()
    all_perms = session.exec(select(RolePermission)).all()
    existing = {(p.role, p.resource): p for p in all_perms}
    for key, val in form.items():
        if not key.startswith("perm_"):
            continue
        parts = key.split("_")
        if len(parts) != 4:
            continue
        _, role, resource, crud = parts
        if crud not in ("c", "r", "u", "d"):
            continue
        perm = existing.get((role, resource))
        if not perm:
            perm = RolePermission(role=role, resource=resource)
            session.add(perm)
            session.flush()
            existing[(role, resource)] = perm
        if crud == "c":
            perm.can_create = str(val).lower() in ("1", "on", "true", "yes")
        elif crud == "r":
            perm.can_read = str(val).lower() in ("1", "on", "true", "yes")
        elif crud == "u":
            perm.can_update = str(val).lower() in ("1", "on", "true", "yes")
        elif crud == "d":
            perm.can_delete = str(val).lower() in ("1", "on", "true", "yes")
        session.add(perm)
    session.commit()
    return RedirectResponse("/users?perms_saved=1", status_code=303)

